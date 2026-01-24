"""
考试管理API

提供考试的CRUD操作
"""

import tempfile
import os
import logging
from pathlib import Path
from typing import Any, Optional, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, and_
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, status, Query, File, UploadFile, Body
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.api.deps import get_db, get_current_active_user
from app.models import User, Exam, UserRole, Semester
from app.models.evaluation import ExamNumberMapping
from app.models.user import User as UserModel
from app.models.organization import School, Classroom
from app.schemas.evaluation import (
    ExamCreate,
    ExamUpdate,
    ExamResponse,
    ExamSubjectCreate,
    ExamSubjectResponse,
)

router = APIRouter()

logger = logging.getLogger(__name__)


# ==================== 考生信息导入相关 Schema ====================

class StudentImportError(BaseModel):
    """学生导入错误"""

    row: int = Field(..., description="行号")
    field: Optional[str] = Field(None, description="字段名")
    message: str = Field(..., description="错误信息")


class StudentImportResponse(BaseModel):
    """学生考生信息批量导入响应"""

    total: int = Field(..., description="总记录数")
    success: int = Field(..., description="成功数")
    failed: int = Field(..., description="失败数")
    created: int = Field(0, description="创建的映射数")
    updated: int = Field(0, description="更新的映射数")
    skipped: int = Field(0, description="跳过的映射数（已存在）")
    errors: List[StudentImportError] = Field(default_factory=list, description="错误列表")


class ExamNumberGenerationResponse(BaseModel):
    """考号生成响应"""

    generated: int = Field(..., description="生成的考号数量")
    conflicts: int = Field(..., description="解决的冲突数量")
    exam_numbers: List[str] = Field(..., description="生成的考号列表（前10个预览）")


@router.post("/", response_model=ExamResponse, status_code=status.HTTP_201_CREATED)
async def create_exam(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    exam_in: ExamCreate,
) -> Any:
    """
    创建新考试

    需要管理员权限
    """
    # 权限检查：只有管理员可以创建考试
    if current_user.role not in [UserRole.ADMIN, UserRole.DISTRICT_ADMIN, UserRole.SCHOOL_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以创建考试"
        )

    # 验证学期是否存在
    semester_result = await db.execute(
        select(Semester).where(Semester.id == exam_in.semester_id)
    )
    semester = semester_result.scalar_one_or_none()
    if not semester:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="学期不存在"
        )

    # 创建考试
    exam = Exam(**exam_in.model_dump(), created_by=current_user.id)
    db.add(exam)
    await db.commit()
    await db.refresh(exam)

    return exam


@router.get("/", response_model=list[ExamResponse])
async def list_exams(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    skip: int = Query(0, ge=0, description="跳过记录数"),
    limit: int = Query(100, ge=1, le=100, description="返回记录数"),
    semester_id: Optional[int] = Query(None, description="学期ID筛选"),
    exam_type: Optional[str] = Query(None, description="考试类型筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    grade_id: Optional[int] = Query(None, description="年级ID筛选"),
    region_id: Optional[int] = Query(None, description="区县ID筛选"),
    school_id: Optional[int] = Query(None, description="学校ID筛选"),
) -> Any:
    """
    获取考试列表

    权限说明：
    - 管理员：可查看所有考试
    - 教研员：可查看所属区县/学校的考试
    - 教师：只能查看所教年级的考试
    - 学生：只能查看自己的考试
    """
    # 构建查询条件
    conditions = []

    # 根据角色限制数据访问
    if current_user.role == UserRole.STUDENT:
        # 学生只能查看自己年级和学校的考试
        if current_user.grade_id:
            conditions.append(Exam.grade_id == current_user.grade_id)
        if current_user.school_id:
            conditions.append(Exam.school_id == current_user.school_id)
    elif current_user.role == UserRole.TEACHER:
        # 教师只能查看所教年级的考试
        # TODO: 根据教师的任教年级筛选
        pass
    # 教研员和管理员可以查看更多数据

    # 应用筛选条件
    if semester_id:
        conditions.append(Exam.semester_id == semester_id)

    if exam_type:
        conditions.append(Exam.exam_type == exam_type)

    if status:
        conditions.append(Exam.status == status)

    if grade_id:
        conditions.append(Exam.grade_id == grade_id)

    if region_id:
        conditions.append(Exam.region_id == region_id)

    if school_id:
        conditions.append(Exam.school_id == school_id)

    # 执行查询
    if conditions:
        query = select(Exam).where(
            and_(*conditions)
        ).order_by(desc(Exam.exam_date)).offset(skip).limit(limit)
    else:
        query = select(Exam).order_by(
            desc(Exam.exam_date)
        ).offset(skip).limit(limit)

    result = await db.execute(query)
    exams = result.scalars().all()

    return exams


@router.get("/{exam_id}", response_model=ExamResponse)
async def get_exam(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    exam_id: int,
) -> Any:
    """
    获取单个考试详情
    """
    result = await db.execute(
        select(Exam).where(Exam.id == exam_id)
    )
    exam = result.scalar_one_or_none()

    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="考试不存在"
        )

    # 权限检查
    # TODO: 根据角色检查数据访问权限

    return exam


@router.put("/{exam_id}", response_model=ExamResponse)
async def update_exam(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    exam_id: int,
    exam_in: ExamUpdate,
) -> Any:
    """
    更新考试信息

    需要管理员权限
    """
    # 权限检查：只有管理员和创建者可以更新考试
    result = await db.execute(
        select(Exam).where(Exam.id == exam_id)
    )
    exam = result.scalar_one_or_none()

    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="考试不存在"
        )

    if current_user.role not in [UserRole.ADMIN, UserRole.DISTRICT_ADMIN, UserRole.SCHOOL_ADMIN]:
        if exam.created_by != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="只有创建者和管理员可以更新考试"
            )

    # 更新字段
    update_data = exam_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(exam, field, value)

    await db.commit()
    await db.refresh(exam)

    return exam


@router.delete("/{exam_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_exam(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    exam_id: int,
) -> None:
    """
    删除考试

    需要管理员权限
    """
    # 权限检查：只有管理员可以删除考试
    if current_user.role not in [UserRole.ADMIN, UserRole.DISTRICT_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有超级管理员和区县管理员可以删除考试"
        )

    result = await db.execute(
        select(Exam).where(Exam.id == exam_id)
    )
    exam = result.scalar_one_or_none()

    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="考试不存在"
        )

    # TODO: 检查是否有关联的成绩，如果有则提示用户先删除成绩

    await db.delete(exam)
    await db.commit()

    return None


@router.post("/{exam_id}/subjects", response_model=ExamSubjectResponse)
async def add_exam_subject(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    exam_id: int,
    subject_in: ExamSubjectCreate,
) -> Any:
    """
    为考试添加科目

    需要管理员权限
    """
    # 权限检查
    if current_user.role not in [UserRole.ADMIN, UserRole.DISTRICT_ADMIN, UserRole.SCHOOL_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以添加考试科目"
        )

    # 验证考试是否存在
    exam_result = await db.execute(
        select(Exam).where(Exam.id == exam_id)
    )
    exam = exam_result.scalar_one_or_none()
    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="考试不存在"
        )

    # 创建考试科目关联
    from app.models import ExamSubject

    exam_subject = ExamSubject(
        exam_id=exam_id,
        **subject_in.model_dump()
    )
    db.add(exam_subject)
    await db.commit()
    await db.refresh(exam_subject)

    return exam_subject


@router.get("/{exam_id}/subjects", response_model=list[ExamSubjectResponse])
async def list_exam_subjects(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    exam_id: int,
) -> Any:
    """
    获取考试的所有科目
    """
    # 验证考试是否存在
    exam_result = await db.execute(
        select(Exam).where(Exam.id == exam_id)
    )
    exam = exam_result.scalar_one_or_none()
    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="考试不存在"
        )

    # 查询考试科目（包含科目名称）
    from app.models import ExamSubject, Subject
    from sqlalchemy.orm import selectinload

    result = await db.execute(
        select(ExamSubject)
        .options(selectinload(ExamSubject.subject))
        .where(ExamSubject.exam_id == exam_id)
        .order_by(ExamSubject.display_order, ExamSubject.id)
    )
    exam_subjects = result.scalars().all()

    # 构建响应，包含科目名称
    response = []
    for exam_subject in exam_subjects:
        # 通过relationship获取subject
        subject_name = exam_subject.subject.name if exam_subject.subject else None
        # 构建响应字典
        subject_dict = {
            'id': exam_subject.id,
            'exam_id': exam_subject.exam_id,
            'subject_id': exam_subject.subject_id,
            'full_score': exam_subject.full_score,
            'pass_line': exam_subject.pass_line,
            'excellent_line': exam_subject.excellent_line,
            'good_line': exam_subject.good_line,
            'display_order': exam_subject.display_order,
            'is_active': exam_subject.is_active,
            'created_at': exam_subject.created_at,
            'subject_name': subject_name,
        }
        response.append(subject_dict)

    return response


# ==================== 考生信息导入 ====================

@router.post("/{exam_id}/students/import", response_model=StudentImportResponse)
async def import_student_exam_info(
    exam_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    批量导入考生信息（考号映射）
    
    Excel格式要求：
    - 必需列：市(区)、学校、姓名、身份证号、考生号、班级
    - 可选列：学校代码
    - 支持格式：.xlsx, .xls
    
    权限说明：
    - 管理员、区县管理员、学校管理员可以导入考生信息
    """
    from app.services.student_import_service import (
        StudentImportService,
        StudentImportServiceError
    )

    # 权限检查
    if current_user.role not in [UserRole.ADMIN, UserRole.DISTRICT_ADMIN, UserRole.SCHOOL_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以导入考生信息"
        )
    
    # 验证考试是否存在
    exam_result = await db.execute(
        select(Exam).where(Exam.id == exam_id)
    )
    exam = exam_result.scalar_one_or_none()
    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="考试不存在"
        )
    
    # 验证文件类型
    if not file.filename:
        logger.error("文件上传失败: 文件名为空")
        raise HTTPException(status_code=400, detail="必须上传文件")
    
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in [".xlsx", ".xls"]:
        logger.error(f"文件上传失败: 不支持的文件格式 {file_ext}, 文件名: {file.filename}")
        raise HTTPException(
            status_code=400,
            detail=f"只支持Excel文件格式 (.xlsx, .xls)，当前文件格式: {file_ext}"
        )
    
    # 保存文件到临时目录
    temp_file_path = None
    try:
        # 创建临时文件
        logger.info(f"开始处理文件上传: {file.filename}, 大小: {file.size if hasattr(file, 'size') else 'unknown'}")
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as temp_file:
            content = await file.read()
            if not content:
                logger.error("文件上传失败: 文件内容为空")
                raise HTTPException(status_code=400, detail="上传的文件为空")
            temp_file.write(content)
            temp_file_path = temp_file.name
            logger.info(f"文件已保存到临时路径: {temp_file_path}, 大小: {len(content)} 字节")
        
        # 解析Excel文件
        try:
            logger.info("开始解析Excel文件...")
            records, parse_errors = await StudentImportService.parse_student_excel(temp_file_path)
            logger.info(f"Excel解析完成: 记录数={len(records)}, 错误数={len(parse_errors)}")
        except StudentImportServiceError as e:
            logger.error(f"Excel解析失败: {str(e)}")
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            logger.error(f"Excel解析异常: {str(e)}", exc_info=True)
            raise HTTPException(status_code=400, detail=f"解析Excel文件失败: {str(e)}")
        
        # 如果解析有错误，返回错误信息
        if parse_errors:
            return StudentImportResponse(
                total=len(records) + len(parse_errors),
                success=0,
                failed=len(parse_errors),
                created=0,
                updated=0,
                skipped=0,
                errors=[StudentImportError(**err) for err in parse_errors]
            )
        
        # 导入考生信息
        result = await StudentImportService.import_student_exam_mappings(
            db, exam_id, records
        )
        
        # 提交事务
        await db.commit()
        
        # 转换错误列表为StudentImportError对象
        error_objects = [StudentImportError(**err) for err in result["errors"]]
        
        logger.info(
            f"考生信息导入完成: 总计={result['total']}, "
            f"成功={result['success']}, 失败={result['failed']}, "
            f"创建={result['created']}, 更新={result['updated']}"
        )
        
        return StudentImportResponse(
            total=result["total"],
            success=result["success"],
            failed=result["failed"],
            created=result["created"],
            updated=result["updated"],
            skipped=result["skipped"],
            errors=error_objects
        )
    
    finally:
        # 清理临时文件
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                logger.info(f"已删除临时文件: {temp_file_path}")
            except Exception as e:
                logger.warning(f"删除临时文件失败: {str(e)}")


# ==================== 考号导出 ====================

@router.get("/{exam_id}/exam-numbers/export")
async def export_exam_numbers(
    exam_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    导出考试考号映射表为Excel

    Excel格式：校级考号 | 市级考号 | 姓名 | 学籍号 | 学校 | 班级

    用途：
    - 区县组织考试时导出校级考号
    - 换算为市级考号后提交给市级考试院
    - 导入市级下发的正式考号
    """
    # 验证考试存在
    exam = await db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="考试不存在"
        )

    logger.info(f"导出考号映射表：考试ID={exam_id}, 考试名称={exam.name}")

    # 查询所有考号映射
    result = await db.execute(
        select(ExamNumberMapping, UserModel, School, Classroom)
        .join(UserModel, ExamNumberMapping.student_id == UserModel.id)
        .outerjoin(School, ExamNumberMapping.school_id == School.id)
        .outerjoin(Classroom, ExamNumberMapping.classroom_id == Classroom.id)
        .where(ExamNumberMapping.exam_id == exam_id)
        .order_by(ExamNumberMapping.exam_number)
    )

    mappings = result.all()

    if not mappings:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="该考试暂无考号映射数据，请先进行考场编排"
        )

    logger.info(f"找到 {len(mappings)} 条考号映射记录")

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "考号映射表"

    # 标题行
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{exam.name} - 考号映射表 - {datetime.now().strftime('%Y-%m-%d')}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 说明文字
    ws.merge_cells('A3:F3')
    ws['A3'] = "说明：此表包含该考试所有学生的考号映射。可导出后换算为市级考号，或导入市级下发的正式考号。"
    ws['A3'].font = Font(size=10, italic=True, color="FF0000")
    ws['A3'].alignment = Alignment(horizontal='left', wrap_text=True, vertical='top')
    ws.row_dimensions[3].height = 30

    # 表头
    headers = ["校级考号", "市级考号", "姓名", "学籍号", "学校", "班级"]
    header_row = 5
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[chr(64 + col_idx)].width = 15

    ws.row_dimensions[header_row].height = 25

    # 数据行
    data_row = 6
    for mapping, student, school, classroom in mappings:
        # 判断考号类型
        exam_number = mapping.exam_number
        if len(exam_number) == 8:
            # 校级考号
            school_exam_number = exam_number
            city_exam_number = ""
        elif len(exam_number) == 10:
            # 市级考号
            school_exam_number = ""
            city_exam_number = exam_number
        else:
            # 其他格式，全部填入校级考号列
            school_exam_number = exam_number
            city_exam_number = ""

        ws.cell(row=data_row, column=1, value=school_exam_number)
        ws.cell(row=data_row, column=2, value=city_exam_number)
        ws.cell(row=data_row, column=3, value=student.full_name or "")
        ws.cell(row=data_row, column=4, value=student.student_id_number or "")
        ws.cell(row=data_row, column=5, value=school.name if school else "")
        ws.cell(row=data_row, column=6, value=f"{classroom.code}班" if classroom else "")

        data_row += 1

    # 保存到临时文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()

    # 生成文件名
    filename = f"{exam.name}_考号映射表_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filename = filename.replace(" ", "_")

    logger.info(f"导出考号映射表成功：{filename}, 共 {len(mappings)} 条记录")

    return FileResponse(
        path=temp_file.name,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==================== 考号生成 ====================

@router.post("/generate-exam-numbers", response_model=ExamNumberGenerationResponse)
async def generate_exam_numbers(
    exam_id: int = Body(..., embed=True, description="考试ID"),
    school_id: int = Body(..., embed=True, description="学校ID"),
    auto_generate: bool = Body(True, embed=True, description="是否自动生成考号"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    批量生成考号

    为指定考试和学校的所有学生生成考号。

    考号格式：school_code (4) + enrollment_year (4) + class_sequence (2) + seat_number (2)
    总长度：12位数字

    权限说明：
    - 管理员、区县管理员、学校管理员可以生成考号

    参数：
    - exam_id: 考试ID
    - school_id: 学校ID
    - auto_generate: 是否自动生成（默认True）

    返回：
    - generated: 生成的考号数量
    - conflicts: 解决的冲突数量
    - exam_numbers: 生成的考号列表（前10个预览）
    """
    from app.utils.exam_number_generator import generate_exam_number, validate_exam_number_async

    # 权限检查
    if current_user.role not in [UserRole.ADMIN, UserRole.DISTRICT_ADMIN, UserRole.SCHOOL_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以生成考号"
        )

    # 验证考试是否存在
    exam = await db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="考试不存在"
        )

    # 验证学校是否存在
    school = await db.get(School, school_id)
    if not school:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="学校不存在"
        )

    logger.info(f"开始批量生成考号：考试ID={exam_id}, 学校ID={school_id}, 学校代码={school.code}")

    # 获取该学校的所有学生
    students_result = await db.execute(
        select(UserModel).where(
            UserModel.school_id == school_id,
            UserModel.role == UserRole.STUDENT
        )
    )
    students = students_result.scalars().all()

    if not students:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"该学校没有学生数据"
        )

    logger.info(f"找到 {len(students)} 名学生")

    generated_numbers = []
    conflicts = 0
    processed = 0
    # Track exam numbers used in this batch to prevent duplicates
    used_exam_numbers = set()

    for student in students:
        # 获取学生班级信息
        if not student.classroom_id:
            logger.warning(f"学生 {student.id} ({student.full_name}) 没有班级信息，跳过")
            continue

        classroom = await db.get(Classroom, student.classroom_id)
        if not classroom:
            logger.warning(f"学生 {student.id} 的班级 {student.classroom_id} 不存在，跳过")
            continue

        # 如果没有入学年份，使用当前年份
        enrollment_year = classroom.enrollment_year if classroom.enrollment_year else datetime.now().year

        # 如果没有班级代码，使用班级ID的后2位
        classroom_code = classroom.code if classroom.code else f"{classroom.id:04d}"

        # 生成考号（座位号默认为1，因为User模型没有seat_number字段）
        # TODO: 未来可以从ExamRoomAssignment获取座位号
        try:
            exam_number = generate_exam_number(
                school_code=school.code,
                enrollment_year=enrollment_year,
                classroom_code=classroom_code,
                seat_number=1  # 默认座位号
            )

            # 验证并处理冲突
            validated_number = await validate_exam_number_async(db, exam_id, exam_number)
            if validated_number != exam_number:
                conflicts += 1
                logger.info(f"考号冲突（数据库）：{exam_number} -> {validated_number}")

            # 检查是否在当前批次中已使用此考号
            if validated_number in used_exam_numbers:
                # 考号在本批次中已被使用，添加后缀
                suffix = 0
                while True:
                    suffix += 1
                    new_number = f"{validated_number}{chr(64 + suffix)}"
                    # 检查数据库和当前批次是否都可用
                    if new_number not in used_exam_numbers:
                        existing = await db.execute(
                            select(ExamNumberMapping).where(
                                ExamNumberMapping.exam_id == exam_id,
                                ExamNumberMapping.exam_number == new_number
                            )
                        )
                        if not existing.scalar_one_or_none():
                            validated_number = new_number
                            conflicts += 1
                            logger.info(f"考号冲突（批次内）：{exam_number} -> {validated_number}")
                            break

            # 标记此考号为已使用
            used_exam_numbers.add(validated_number)

            # 创建考号映射
            # 检查是否已存在
            existing_mapping = await db.execute(
                select(ExamNumberMapping).where(
                    ExamNumberMapping.exam_id == exam_id,
                    ExamNumberMapping.student_id == student.id
                )
            )
            existing = existing_mapping.scalar_one_or_none()

            if existing:
                # 更新现有映射
                existing.exam_number = validated_number
                existing.school_id = school_id
                existing.classroom_id = classroom.id
            else:
                # 创建新映射前，再次检查考号是否已被其他学生使用
                existing_by_number_result = await db.execute(
                    select(ExamNumberMapping).where(
                        ExamNumberMapping.exam_id == exam_id,
                        ExamNumberMapping.exam_number == validated_number
                    )
                )
                existing_by_number = existing_by_number_result.scalar_one_or_none()

                if existing_by_number:
                    # 考号已被使用，记录警告并跳过
                    logger.warning(
                        f"考号 {validated_number} 已被学生 {existing_by_number.student_id} 使用，"
                        f"跳过学生 {student.id}"
                    )
                    continue

                # 创建新映射
                mapping = ExamNumberMapping(
                    exam_id=exam_id,
                    exam_number=validated_number,
                    student_id=student.id,
                    student_id_number=student.student_id_number or "",
                    school_id=school_id,
                    classroom_id=classroom.id
                )
                db.add(mapping)

            generated_numbers.append(validated_number)
            processed += 1

        except ValueError as e:
            logger.error(f"生成考号失败：学生 {student.id}, 错误：{str(e)}")
            continue
        except Exception as e:
            # 捕获其他异常（如唯一约束冲突）
            if "duplicate key" in str(e) or "UniqueViolation" in str(e):
                logger.warning(f"考号冲突：{validated_number}，尝试添加后缀")
                # 添加后缀重试
                suffix = 0
                while True:
                    suffix += 1
                    new_number = f"{validated_number}{chr(64 + suffix)}"

                    # 检查新考号是否可用
                    existing = await db.execute(
                        select(ExamNumberMapping).where(
                            ExamNumberMapping.exam_id == exam_id,
                            ExamNumberMapping.exam_number == new_number
                        )
                    ).scalar_one_or_none()

                    if not existing:
                        # 创建新映射
                        mapping = ExamNumberMapping(
                            exam_id=exam_id,
                            exam_number=new_number,
                            student_id=student.id,
                            student_id_number=student.student_id_number or "",
                            school_id=school_id,
                            classroom_id=classroom.id
                        )
                        db.add(mapping)
                        generated_numbers.append(new_number)
                        conflicts += 1
                        processed += 1
                        logger.info(f"解决冲突：{validated_number} -> {new_number}")
                        break
            else:
                logger.error(f"生成考号失败：学生 {student.id}, 错误：{str(e)}")
                continue

    # 提交所有更改
    await db.commit()

    logger.info(
        f"考号生成完成：处理={processed}, 生成={len(generated_numbers)}, "
        f"冲突={conflicts}, 学生总数={len(students)}"
    )

    return ExamNumberGenerationResponse(
        generated=len(generated_numbers),
        conflicts=conflicts,
        exam_numbers=generated_numbers[:10]  # 返回前10个作为预览
    )
