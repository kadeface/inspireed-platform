"""
Unified Import API Endpoint

Provides a single endpoint for all Excel-based imports.
"""

import tempfile
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.database import get_db
from app.api.deps import get_current_admin
from app.models import User
from app.schemas.import_schemas import (
    ImportStrategyType,
    ImportResult,
    SchoolImportResult,
    TeacherImportResult,
    SchoolImportResponse,
    SchoolImportError,
)
from app.services.import_orchestrator import ImportOrchestrator

import logging

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/import", tags=["import"])


@router.post("", response_model=ImportResult)
async def unified_import(
    strategy_type: ImportStrategyType = Query(..., description="Import strategy type"),
    file: UploadFile = File(..., description="Excel file"),
    update_existing: bool = Query(False, description="Update existing records"),
    auto_create: bool = Query(True, description="Auto-create related entities (regions, teachers, etc.)"),
    exam_id: Optional[int] = Query(None, description="Exam ID (for student imports)"),
    school_id: Optional[int] = Query(None, description="School ID (for classroom imports)"),
    region_id: Optional[int] = Query(None, description="Region ID (for classroom imports)"),
    is_school_admin: bool = Query(False, description="School admin mode (for classroom imports)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    """
    Unified import endpoint for Excel-based imports

    Supports:
    - **school**: Import schools with region auto-creation
    - **classroom**: Import classrooms (dual-mode: district or school admin)
    - **student**: Import student exam mappings (requires exam_id)
    - **teacher**: Import teacher assignments (with teacher/semester auto-creation)
    - **city_exam_number**: Import city exam numbers (requires exam_id)

    **Parameters:**
    - **strategy_type**: Import type (school, classroom, student, teacher, city_exam_number)
    - **file**: Excel file (.xlsx or .xls)
    - **update_existing**: Update existing records instead of skipping them
    - **auto_create**: Auto-create related entities (regions, teachers, semesters)
    - **exam_id**: Required for student and city_exam_number imports
    - **school_id**: Required for classroom imports in school admin mode
    - **region_id**: Optional region ID for classroom imports
    - **is_school_admin**: Enable school admin mode for classroom imports

    **Returns:**
    Import result with statistics and error details
    """

    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="必须上传文件")

    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in [".xlsx", ".xls"]:
        raise HTTPException(
            status_code=400,
            detail=f"只支持Excel文件格式 (.xlsx, .xls)，当前文件格式: {file_ext}"
        )

    # Build context
    context = {
        "update_existing": update_existing,
        "auto_create": auto_create,
        "auto_create_region": auto_create,
        "auto_create_teachers": auto_create,
        "auto_create_semesters": auto_create,
        "exam_id": exam_id,
        "school_id": school_id,
        "region_id": region_id,
        "is_school_admin": is_school_admin,
    }

    # Save to temp file
    temp_file_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as temp_file:
            content = await file.read()
            if not content:
                raise HTTPException(status_code=400, detail="上传的文件为空")
            temp_file.write(content)
            temp_file_path = Path(temp_file.name)

        logger.info(
            f"Processing {strategy_type} import: {file.filename} "
            f"({len(content)} bytes)"
        )

        # Execute import
        orchestrator = ImportOrchestrator()
        result = await orchestrator.execute_import(
            db=db,
            strategy_type=strategy_type.value,
            file_path=temp_file_path,
            context=context
        )

        # Return appropriate response type based on strategy
        if strategy_type == ImportStrategyType.SCHOOL:
            return SchoolImportResult(**result)
        elif strategy_type == ImportStrategyType.TEACHER:
            return TeacherImportResult(**result)
        else:
            return ImportResult(**result)

    except HTTPException:
        raise
    except ValueError as e:
        # Invalid strategy type
        logger.error(f"Invalid strategy: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Import failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")

    finally:
        # Clean up temp file
        if temp_file_path and temp_file_path.exists():
            try:
                temp_file_path.unlink()
            except Exception as e:
                logger.warning(f"Failed to delete temp file: {str(e)}")


@router.get("/strategies")
async def list_strategies():
    """List available import strategies

    Returns information about supported import types and their requirements.
    """
    return {
        "strategies": [
            {
                "type": "school",
                "name": "学校导入",
                "description": "导入学校信息，支持自动创建区域",
                "required_context": [],
                "optional_context": ["auto_create_region"],
            },
            {
                "type": "classroom",
                "name": "班级导入",
                "description": "导入班级信息，支持县区端和学校端两种模式",
                "required_context": [],
                "optional_context": ["school_id", "region_id", "is_school_admin", "update_existing"],
            },
            {
                "type": "student",
                "name": "学生考号导入",
                "description": "导入学生考试映射，将学生与考号关联",
                "required_context": ["exam_id"],
                "optional_context": [],
            },
            {
                "type": "student_account",
                "name": "学生账户导入",
                "description": "批量创建学生账户，格式：学校名称、年级级别、班级编号、学籍号、姓名",
                "required_context": [],
                "optional_context": ["school_id", "region_id", "is_school_admin", "update_existing"],
            },
            {
                "type": "teacher",
                "name": "教师导入",
                "description": "导入教师教学任务，支持自动创建教师和学期",
                "required_context": [],
                "optional_context": ["auto_create_teachers", "auto_create_semesters", "update_existing"],
            },
            {
                "type": "city_exam_number",
                "name": "市级考号导入",
                "description": "导入市级考试院下发的考号映射表，格式：校级考号、市级考号、姓名、学校",
                "required_context": ["exam_id"],
                "optional_context": ["update_existing"],
            },
        ]
    }


@router.get("/template/{strategy_type}")
async def download_import_template(
    strategy_type: ImportStrategyType,
    is_school_admin: bool = Query(False, description="School admin mode (for classroom/student_account templates)"),
    current_user: User = Depends(get_current_admin),
):
    """
    Download Excel import template for specified strategy type

    Generates and returns an Excel template file with proper column headers and example data.
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"

    # Define template configurations
    templates = {
        ImportStrategyType.SCHOOL: {
            "name": "学校导入模板",
            "filename": "school_import_template.xlsx",
            "headers": ["学校名称*", "学校代码", "区域名称"],
            "examples": ["第一中学", "101", "朝阳区"],
            "descriptions": ["必填", "可选", "可选"],
        },
        ImportStrategyType.CLASSROOM: {
            "name": "班级导入模板",
            "filename": "classroom_import_template.xlsx",
            "headers_district": ["学校名称*", "学校代码", "年级级别*", "年级名称", "班级编号*", "班级名称", "入学年份", "班级容量"],
            "examples_district": ["第一中学", "101", "10", "高一", "1001", "高一1班", "2024", "50"],
            "headers_school": ["年级级别*", "年级名称", "班级编号*", "班级名称", "入学年份", "班级容量"],
            "examples_school": ["10", "高一", "1001", "高一1班", "2024", "50"],
            "descriptions": ["必填", "可选", "必填", "可选", "可选", "可选"],
        },
        ImportStrategyType.STUDENT: {
            "name": "学生考号导入模板",
            "filename": "student_exam_import_template.xlsx",
            "headers": ["市(区)", "学校", "学校代码", "姓名*", "身份证号*", "考生号*", "班级"],
            "examples": ["朝阳区", "第一中学", "101", "张三", "110101202401011234", "202410001", "1001"],
            "descriptions": ["必填", "必填", "可选", "必填", "必填", "必填", "必填"],
        },
        ImportStrategyType.STUDENT_ACCOUNT: {
            "name": "学生账户导入模板",
            "filename": "student_account_import_template.xlsx",
            "headers_district": ["学校名称*", "学校代码", "年级级别*", "班级编号*", "学籍号*", "姓名*", "用户名", "邮箱", "手机号", "性别"],
            "examples_district": ["第一中学", "101", "10", "1001", "2024100001", "张三", "2024100001", "2024100001@inspireed.com", "13800138000", "男"],
            "headers_school": ["年级级别*", "班级编号*", "学籍号*", "姓名*", "用户名", "邮箱", "手机号", "性别"],
            "examples_school": ["10", "1001", "2024100001", "张三", "2024100001", "2024100001@inspireed.com", "13800138000", "男"],
            "descriptions": ["必填", "可选", "必填", "必填", "必填", "必填", "可选", "可选", "可选", "可选"],
        },
        ImportStrategyType.TEACHER: {
            "name": "教师导入模板",
            "filename": "teacher_import_template.xlsx",
            "headers": ["学校名称*", "学校代码", "学期名称*", "年级级别*", "班级编号*", "科目*", "教师姓名*", "教师工号", "教师类型"],
            "examples": ["第一中学", "101", "2024-2025第一学期", "10", "1001", "数学", "李老师", "T001", "主讲教师"],
            "descriptions": ["必填", "可选", "必填", "必填", "必填", "必填", "必填", "可选", "可选"],
        },
        ImportStrategyType.CITY_EXAM_NUMBER: {
            "name": "市级考号导入模板",
            "filename": "city_exam_number_import_template.xlsx",
            "headers": ["校级考号*", "市级考号*", "姓名*", "学校"],
            "examples": ["20250101", "7783190101", "张三", "XX中学"],
            "descriptions": ["必填", "必填", "必填", "可选"],
        },
    }

    # Get template config
    template = templates.get(strategy_type)
    if not template:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown strategy type: {strategy_type}"
        )

    # Determine headers and examples based on mode
    if strategy_type == ImportStrategyType.CLASSROOM:
        if is_school_admin:
            headers = template["headers_school"]
            examples = template["examples_school"]
        else:
            headers = template["headers_district"]
            examples = template["examples_district"]
    elif strategy_type == ImportStrategyType.STUDENT_ACCOUNT:
        if is_school_admin:
            headers = template["headers_school"]
            examples = template["examples_school"]
        else:
            headers = template["headers_district"]
            examples = template["examples_district"]
    else:
        headers = template["headers"]
        examples = template["examples"]

    descriptions = template.get("descriptions", [""] * len(headers))

    # Title row
    ws.merge_cells('A1:Z1')
    ws['A1'] = f"{template['name']} - {datetime.now().strftime('%Y-%m-%d')}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    # Instructions
    ws.merge_cells('A3:Z3')
    ws['A3'] = "说明：请严格按照表头填写数据，带*号的为必填项。填写完成后删除示例数据行。"
    ws['A3'].font = Font(size=10, italic=True, color="FF0000")
    ws['A3'].alignment = Alignment(horizontal='left', wrap_text=True)
    ws.row_dimensions[3].height = 20

    # Header row
    header_row = 5
    for col_idx, (header, desc) in enumerate(zip(headers, descriptions), 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[chr(64 + col_idx)].width = 15

    ws.row_dimensions[header_row].height = 25

    # Example row
    example_row = 6
    for col_idx, example in enumerate(examples, 1):
        cell = ws.cell(row=example_row, column=col_idx)
        cell.value = example
        cell.font = Font(italic=True, color="666666")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Description row
    desc_row = 7
    for col_idx, desc in enumerate(descriptions, 1):
        cell = ws.cell(row=desc_row, column=col_idx)
        cell.value = desc
        cell.font = Font(size=9, color="999999")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add more example rows for student_account
    if strategy_type == ImportStrategyType.STUDENT_ACCOUNT:
        # Add 3 more example rows
        additional_examples = [
            ["第一中学", "101", "10", "1001", "2024100002", "李四", "2024100002", "2024100002@inspireed.com", "13800138001", "女"] if not is_school_admin else ["10", "1001", "2024100002", "李四", "2024100002", "2024100002@inspireed.com", "13800138001", "女"],
            ["第一中学", "101", "10", "1002", "2024100003", "王五", "2024100003", "2024100003@inspireed.com", "13800138002", "男"] if not is_school_admin else ["10", "1002", "2024100003", "王五", "2024100003", "2024100003@inspireed.com", "13800138002", "男"],
        ]

        for row_offset, example_data in enumerate(additional_examples, 1):
            for col_idx, value in enumerate(example_data, 1):
                cell = ws.cell(row=example_row + row_offset, column=col_idx)
                cell.value = value
                cell.font = Font(italic=True, color="666666")
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save to temp file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()

    # Return file
    return FileResponse(
        path=temp_file.name,
        filename=template['filename'],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
