"""
Excel成绩导入服务

提供Excel文件解析、数据验证和批量导入功能
"""

import logging
from typing import Any, Awaitable, Callable, Dict, List, Optional, Tuple
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from sqlalchemy.exc import IntegrityError

from app.models import (
    User,
    Exam,
    Subject,
    Score,
    ImportStatus,
    ImportTask,
    ExamNumberMapping,
    ExamSubject,
    UserRole,
)

logger = logging.getLogger(__name__)


class ExcelImportError(Exception):
    """Excel导入错误"""
    pass


class ExcelImportService:
    """Excel成绩导入服务"""

    # Excel列名映射（支持多种格式）
    COLUMN_MAPPING = {
        # 考号（多种命名方式）
        "考号": "exam_number",
        "考试号": "exam_number",
        "准考证号": "exam_number",
        # 学籍号
        "学籍号": "student_id_number",
        "学生学籍号": "student_id_number",
        "身份证号": "student_id_number",
        # 姓名
        "姓名": "name",
        "学生姓名": "name",
        # 科目
        "科目": "subject_name",
        "学科": "subject_name",
        "考试科目": "subject_name",
        # 原始分
        "原始分": "raw_score",
        "分数": "raw_score",
        "成绩": "raw_score",
        "得分": "raw_score",
        # 缺考标记
        "缺考": "is_absent",
        "是否缺考": "is_absent",
        # 作弊标记
        "作弊": "is_cheated",
        "是否作弊": "is_cheated",
    }

    # 必需列
    REQUIRED_COLUMNS = ["考号", "学籍号", "科目", "原始分"]

    @staticmethod
    async def parse_excel_file(file_path: str) -> List[Dict[str, Any]]:
        """
        解析Excel文件

        Args:
            file_path: Excel文件路径

        Returns:
            成绩记录列表

        Raises:
            ExcelImportError: 文件格式错误或数据不完整
        """
        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise ExcelImportError("Excel文件为空或没有工作表")

            # 获取表头
            headers = [cell.value for cell in ws[1]]  # type: ignore
            if not headers:
                raise ExcelImportError("Excel文件为空或没有表头")

            # 标准化列名（去除空格）
            headers = [str(h).strip() if h else "" for h in headers]

            # 查找必需列
            column_indices = {}
            for col_name, field_name in ExcelImportService.COLUMN_MAPPING.items():
                if col_name in headers:
                    column_indices[field_name] = headers.index(col_name)

            # 验证必需列
            required_fields = ["exam_number", "student_id_number", "subject_name", "raw_score"]
            missing_fields = [f for f in required_fields if f not in column_indices]
            if missing_fields:
                raise ExcelImportError(f"缺少必需列: {missing_fields}")

            # 解析数据行
            records = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # type: ignore
                if not any(row):  # 跳过空行
                    continue

                try:
                    record: Dict[str, Any] = {
                        "row_number": row_idx,
                    }

                    # 提取考号
                    exam_number = row[column_indices["exam_number"]]  # type: ignore
                    if exam_number is None:
                        continue
                    record["exam_number"] = str(exam_number).strip()

                    # 提取学籍号
                    student_id_number = row[column_indices["student_id_number"]]  # type: ignore
                    if student_id_number is None:
                        continue
                    record["student_id_number"] = str(student_id_number).strip()

                    # 提取姓名（可选）
                    if "name" in column_indices:
                        name = row[column_indices["name"]]  # type: ignore
                        record["name"] = str(name).strip() if name else ""

                    # 提取科目
                    subject_name = row[column_indices["subject_name"]]  # type: ignore
                    if subject_name is None:
                        continue
                    record["subject_name"] = str(subject_name).strip()

                    # 提取原始分
                    raw_score = row[column_indices["raw_score"]]  # type: ignore
                    if raw_score is None:
                        # 检查是否缺考
                        is_absent = False
                        if "is_absent" in column_indices:
                            absent_value = row[column_indices["is_absent"]]  # type: ignore
                            is_absent = absent_value in ["是", "Y", "yes", "1", True, 1]

                        # 检查是否作弊
                        is_cheated = False
                        if "is_cheated" in column_indices:
                            cheated_value = row[column_indices["is_cheated"]]  # type: ignore
                            is_cheated = cheated_value in ["是", "Y", "yes", "1", True, 1]

                        if not is_absent and not is_cheated:
                            continue  # 没有分数也不是缺考/作弊，跳过

                        record["raw_score"] = 0
                        record["is_absent"] = is_absent
                        record["is_cheated"] = is_cheated
                    else:
                        try:
                            record["raw_score"] = float(raw_score)  # type: ignore
                            record["is_absent"] = False
                            record["is_cheated"] = False
                        except (ValueError, TypeError):
                            raise ExcelImportError(f"第{row_idx}行: 分数格式错误 '{raw_score}'")

                    records.append(record)

                except Exception as e:
                    logger.warning(f"解析第{row_idx}行失败: {e}")
                    raise ExcelImportError(f"第{row_idx}行: {str(e)}")

            wb.close()
            return records

        except Exception as e:
            if isinstance(e, ExcelImportError):
                raise
            raise ExcelImportError(f"解析Excel文件失败: {str(e)}")

    @staticmethod
    async def validate_record(
        db: AsyncSession,
        record: Dict[str, Any],
        exam_id: int,
    ) -> Tuple[bool, Optional[str], Optional[Dict[str, Any]]]:
        """
        验证单条成绩记录

        Args:
            db: 数据库会话
            record: 成绩记录
            exam_id: 考试ID

        Returns:
            (是否有效, 错误信息, 验证后的数据)
        """
        # 验证考号映射
        mapping_result = await db.execute(
            select(ExamNumberMapping).where(
                and_(
                    ExamNumberMapping.exam_id == exam_id,
                    ExamNumberMapping.exam_number == record["exam_number"]
                )
            )
        )
        mapping = mapping_result.scalar_one_or_none()
        if not mapping:
            return False, f"考号 '{record['exam_number']}' 不存在或未关联到该考试", None

        # 验证学籍号
        student_result = await db.execute(
            select(User).where(User.student_id_number == record["student_id_number"])
        )
        student = student_result.scalar_one_or_none()
        if not student:
            return False, f"学籍号 '{record['student_id_number']}' 不存在", None

        # 验证考号和学籍号是否匹配
        if mapping.student_id != student.id:  # type: ignore
            return False, f"考号 '{record['exam_number']}' 与学籍号 '{record['student_id_number']}' 不匹配", None

        # 验证科目：通过exam_id和subject_name查找ExamSubject
        # 需要join Subject表来获取科目名称
        exam_subject_result = await db.execute(
            select(ExamSubject)
            .join(Subject, ExamSubject.subject_id == Subject.id)
            .where(
                and_(
                    ExamSubject.exam_id == exam_id,
                    Subject.name == record["subject_name"],
                    ExamSubject.is_active == True
                )
            )
        )
        exam_subject = exam_subject_result.scalar_one_or_none()
        if not exam_subject:
            # 检查科目是否存在（但不在此考试中）
            subject_result = await db.execute(
                select(Subject).where(Subject.name == record["subject_name"])
            )
            subject = subject_result.scalar_one_or_none()
            if subject:
                return False, f"科目 '{record['subject_name']}' 不在该考试的科目列表中，请先在考试中添加该科目", None
            else:
                return False, f"科目 '{record['subject_name']}' 不存在于系统中", None

        # 获取科目信息（用于后续验证）
        subject = exam_subject.subject

        # 验证分数范围（使用ExamSubject中的full_score作为上限）
        if not record["is_absent"] and not record["is_cheated"]:
            raw_score = record["raw_score"]
            max_score = exam_subject.full_score or 150
            if raw_score < 0 or raw_score > max_score:
                return False, f"分数 {raw_score} 超出合理范围 (0-{max_score})，该科目满分为 {max_score}", None

        # 组装验证后的数据
        validated_data = {
            "exam_id": exam_id,
            "subject_id": exam_subject.subject_id,  # 使用ExamSubject中的subject_id
            "student_id": student.id,
            "raw_score": record["raw_score"],
            "is_absent": record["is_absent"],
            "is_cheated": record["is_cheated"],
            "standard_score": None,  # 待计算
            "percentile": None,  # 待计算
            "grade_level": None,  # 待计算（可根据exam_subject中的分数线计算）
        }

        return True, None, validated_data

    @staticmethod
    async def import_scores(
        db: AsyncSession,
        exam_id: int,
        records: List[Dict[str, Any]],
        task_id: Optional[int] = None,
        progress_callback: Optional[Callable[[int, int, int, int, int], Awaitable[None]]] = None,
    ) -> Dict[str, Any]:
        """
        批量导入成绩

        Args:
            db: 数据库会话
            exam_id: 考试ID
            records: 成绩记录列表
            task_id: 导入任务ID（用于进度更新）
            progress_callback: 进度回调函数

        Returns:
            导入结果统计
        """
        total = len(records)
        success_count = 0
        failed_count = 0
        errors = []

        for idx, record in enumerate(records):
            try:
                # 验证记录
                is_valid, error_msg, validated_data = await ExcelImportService.validate_record(
                    db, record, exam_id
                )

                if not is_valid or validated_data is None:
                    failed_count += 1
                    errors.append({
                        "row": record.get("row_number", "unknown"),
                        "exam_number": record.get("exam_number", "unknown"),
                        "error": error_msg or "验证失败",
                    })
                    logger.warning(f"验证失败: {error_msg}")
                    continue

                # 检查是否已存在（避免重复导入）
                student_id = validated_data.get("student_id")  # type: ignore
                subject_id = validated_data.get("subject_id")  # type: ignore
                if student_id is None or subject_id is None:
                    failed_count += 1
                    errors.append({
                        "row": record.get("row_number", "unknown"),
                        "exam_number": record.get("exam_number", "unknown"),
                        "error": "验证后的数据缺少必要字段",
                    })
                    continue

                existing_result = await db.execute(
                    select(Score).where(
                        and_(
                            Score.exam_id == exam_id,
                            Score.student_id == student_id,
                            Score.subject_id == subject_id,
                        )
                    )
                )
                existing_score = existing_result.scalar_one_or_none()

                if existing_score:
                    # 更新现有记录
                    existing_score.raw_score = validated_data.get("raw_score", 0)  # type: ignore
                    existing_score.is_absent = validated_data.get("is_absent", False)  # type: ignore
                    existing_score.is_cheated = validated_data.get("is_cheated", False)  # type: ignore
                    logger.info(f"更新成绩: 学生{student_id} 科目{subject_id}")
                else:
                    # 创建新记录
                    new_score = Score(**validated_data)  # type: ignore
                    db.add(new_score)
                    logger.info(f"创建成绩: 学生{student_id} 科目{subject_id}")

                success_count += 1

                # 每处理100条提交一次
                if (idx + 1) % 100 == 0:
                    await db.commit()
                    logger.info(f"已处理 {idx + 1}/{total} 条记录")

                # 更新进度
                if progress_callback and task_id is not None:
                    progress = int((idx + 1) / total * 100)
                    await progress_callback(task_id, progress, idx + 1, success_count, failed_count)

            except Exception as e:
                failed_count += 1
                error_msg = f"处理失败: {str(e)}"
                errors.append({
                    "row": record.get("row_number", "unknown"),
                    "exam_number": record.get("exam_number", "unknown"),
                    "error": error_msg,
                })
                logger.error(f"导入失败: {error_msg}")

        # 最终提交
        await db.commit()

        return {
            "total": total,
            "success": success_count,
            "failed": failed_count,
            "errors": errors[:100],  # 只返回前100个错误
        }

    @staticmethod
    async def process_import_task(
        db: AsyncSession,
        task_id: int,
    ) -> Dict[str, Any]:
        """
        处理导入任务（异步）

        Args:
            db: 数据库会话
            task_id: 导入任务ID

        Returns:
            处理结果
        """
        # 获取任务
        task_result = await db.execute(
            select(ImportTask).where(ImportTask.id == task_id)
        )
        task = task_result.scalar_one_or_none()
        if not task:
            raise ExcelImportError(f"导入任务 {task_id} 不存在")

        # 更新任务状态
        task.status = ImportStatus.PROCESSING  # type: ignore
        task.started_at = datetime.utcnow()  # type: ignore
        await db.commit()

        try:
            # 解析Excel文件
            file_path = task.file_url.replace("file://", "")  # 移除URL前缀
            records = await ExcelImportService.parse_excel_file(file_path)

            # 更新任务信息
            task.total_rows = len(records)  # type: ignore
            task.progress = 10  # type: ignore
            await db.commit()

            # 导入成绩
            async def progress_callback(task_id: int, progress: int, processed: int, success: int, failed: int) -> None:
                task_result = await db.execute(
                    select(ImportTask).where(ImportTask.id == task_id)
                )
                task = task_result.scalar_one_or_none()
                if task:
                    task.progress = 10 + int(progress * 0.9)  # 10%-100%  # type: ignore
                    task.processed_rows = processed  # type: ignore
                    await db.commit()

            exam_id = int(task.exam_id)  # type: ignore
            result = await ExcelImportService.import_scores(
                db,
                exam_id,
                records,
                task_id,
                progress_callback,
            )

            # 更新任务状态
            task.status = ImportStatus.COMPLETED  # type: ignore
            task.progress = 100  # type: ignore
            task.processed_rows = result["success"] + result["failed"]  # type: ignore
            task.failed_rows = result["failed"]  # type: ignore
            task.completed_at = datetime.utcnow()  # type: ignore
            if result["errors"]:
                task.error_message = f"导入完成，但有 {result['failed']} 条记录失败"  # type: ignore
                task.error_details = {"errors": result["errors"]}  # type: ignore
            await db.commit()

            logger.info(f"导入任务 {task_id} 完成: 成功 {result['success']}, 失败 {result['failed']}")
            return result

        except Exception as e:
            # 更新任务状态为失败
            task.status = ImportStatus.FAILED  # type: ignore
            task.error_message = str(e)  # type: ignore
            task.completed_at = datetime.utcnow()  # type: ignore
            await db.commit()

            logger.error(f"导入任务 {task_id} 失败: {e}")
            raise
