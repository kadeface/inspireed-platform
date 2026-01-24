"""
教师教学任务批量导入服务

提供Excel文件解析、数据验证和批量导入功能
"""

import logging
import re
from typing import Any, Dict, List, Optional, Tuple
from pathlib import Path
from datetime import datetime, date

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_
from sqlalchemy.exc import IntegrityError
from pypinyin import lazy_pinyin, Style

from app.models import (
    User,
    UserRole,
    School,
    Grade,
    Classroom,
    Subject,
    Semester,
    TeacherTeachingAssignment,
    TeachingAssignmentType,
    TeacherPositionType,
)
from app.core.security import get_password_hash

logger = logging.getLogger(__name__)


class TeacherAssignmentImportServiceError(Exception):
    """教师教学任务导入服务错误"""
    pass


class TeacherAssignmentImportService:
    """教师教学任务批量导入服务"""

    # Excel列名映射（支持多种格式）
    COLUMN_MAPPING = {
        # 教师姓名
        "教师姓名": "teacher_name",
        "教师": "teacher_name",
        "教师姓名*": "teacher_name",
        "姓名": "teacher_name",
        # 学校名称
        "学校名称": "school_name",
        "学校": "school_name",
        "学校名称*": "school_name",
        # 年级名称/年级级别
        "年级名称": "grade_name",
        "年级": "grade_name",
        "年级名称*": "grade_name",
        "年级级别": "grade_level",
        "年级级别*": "grade_level",
        "级别": "grade_level",
        # 班级名称/班级编码
        "班级名称": "classroom_name",
        "班级": "classroom_name",
        "班级名称*": "classroom_name",
        "班级编码": "classroom_code",
        "班级编码*": "classroom_code",
        "班级编号": "classroom_code",
        "班级代码": "classroom_code",
        # 学科名称
        "学科名称": "subject_name",
        "学科": "subject_name",
        "学科名称*": "subject_name",
        # 学期名称/学期编号
        "学期名称": "semester_name",
        "学期": "semester_name",
        "学期名称*": "semester_name",
        "学期编号": "semester_number",
        "学期编号*": "semester_number",
        "学期": "semester_number",
        # 学年
        "学年": "academic_year",
        "学年*": "academic_year",
        # 任务类型
        "任务类型": "assignment_type",
        "任务类型*": "assignment_type",
        "类型": "assignment_type",
        # 是否激活
        "是否激活": "is_active",
        "激活": "is_active",
        "状态": "is_active",
    }

    @staticmethod
    def parse_assignment_excel(file_path: Path) -> List[Dict[str, Any]]:
        """
        解析教师教学任务Excel文件

        Args:
            file_path: Excel文件路径

        Returns:
            解析后的数据列表，每个元素是一个字典，包含字段名和值

        Raises:
            TeacherAssignmentImportServiceError: 如果文件格式不正确
        """
        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active

            if worksheet is None:
                raise TeacherAssignmentImportServiceError("Excel文件格式错误：无法读取工作表")

            # 读取表头
            headers = []
            header_row = None
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                if row and any(cell for cell in row if cell):
                    # 找到第一个非空行作为表头
                    headers = [str(cell).strip() if cell else "" for cell in row]
                    header_row = row_idx
                    break

            if not headers or header_row is None:
                raise TeacherAssignmentImportServiceError("Excel文件格式错误：未找到表头")

            # 映射列名
            column_index_map = {}
            for idx, header in enumerate(headers):
                header_clean = header.strip()
                if header_clean in TeacherAssignmentImportService.COLUMN_MAPPING:
                    field_name = TeacherAssignmentImportService.COLUMN_MAPPING[header_clean]
                    column_index_map[field_name] = idx

            # 检查必填字段（支持新旧格式）
            required_fields_grade = ["grade_name", "grade_level"]  # 年级名称或年级级别至少一个
            required_fields_classroom = ["classroom_name", "classroom_code"]  # 班级名称或班级编码至少一个
            required_fields_semester = ["semester_name", "semester_number"]  # 学期名称或学期编号至少一个
            required_fields_base = ["teacher_name", "school_name", "subject_name", "academic_year", "assignment_type"]
            
            # 检查基础必填字段
            missing_base = [field for field in required_fields_base if field not in column_index_map]
            if missing_base:
                raise TeacherAssignmentImportServiceError(
                    f"Excel文件缺少必填列：{', '.join(missing_base)}"
                )
            
            # 检查年级字段（至少有一个）
            if not any(field in column_index_map for field in required_fields_grade):
                raise TeacherAssignmentImportServiceError(
                    f"Excel文件缺少年级字段（需要：年级名称 或 年级级别）"
                )
            
            # 检查班级字段（至少有一个）
            if not any(field in column_index_map for field in required_fields_classroom):
                raise TeacherAssignmentImportServiceError(
                    f"Excel文件缺少班级字段（需要：班级名称 或 班级编码）"
                )
            
            # 检查学期字段（至少有一个）
            if not any(field in column_index_map for field in required_fields_semester):
                raise TeacherAssignmentImportServiceError(
                    f"Excel文件缺少学期字段（需要：学期名称 或 学期编号）"
                )

            # 读取数据行
            data = []
            for row_idx, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
            ):
                # 跳过空行
                if not any(cell for cell in row if cell):
                    continue

                row_data = {}
                for field_name, col_idx in column_index_map.items():
                    value = row[col_idx] if col_idx < len(row) else None
                    if value is not None:
                        # 处理字符串类型
                        if isinstance(value, str):
                            value = value.strip()
                            if value == "":
                                value = None
                        row_data[field_name] = value

                # 只添加至少有一个字段的行
                if row_data:
                    row_data["_row_number"] = row_idx
                    data.append(row_data)

            if not data:
                raise TeacherAssignmentImportServiceError("Excel文件中没有有效数据")

            return data

        except Exception as e:
            if isinstance(e, TeacherAssignmentImportServiceError):
                raise
            raise TeacherAssignmentImportServiceError(f"解析Excel文件失败：{str(e)}")

    @staticmethod
    async def find_teacher(
        db: AsyncSession,
        teacher_name: str,
    ) -> Optional[User]:
        """
        根据教师姓名查找教师

        Args:
            db: 数据库会话
            teacher_name: 教师姓名

        Returns:
            教师对象，如果未找到返回None
        """
        if not teacher_name:
            return None

        teacher_name = str(teacher_name).strip()

        # 尝试通过姓名匹配
        result = await db.execute(
            select(User).where(
                and_(
                    User.role == UserRole.TEACHER,
                    or_(
                        User.full_name == teacher_name,
                        User.username == teacher_name,
                        User.email == teacher_name,
                    ),
                )
            )
        )
        teacher = result.scalar_one_or_none()

        return teacher

    @staticmethod
    def _name_to_pinyin(name: str) -> str:
        """
        将中文姓名转换为拼音（小写，无空格）

        Args:
            name: 中文姓名

        Returns:
            拼音字符串，如 "zhanglaoshi"
        """
        if not name:
            return ""
        # 使用 lazy_pinyin 转换为拼音，然后连接
        pinyin_list = lazy_pinyin(name, style=Style.NORMAL)
        return "".join(pinyin_list).lower()

    @staticmethod
    def _generate_username(
        teacher_name: str,
        school_code: str,
        classroom_code: Optional[str] = None,
        is_duplicate: bool = False,
    ) -> str:
        """
        生成教师用户名

        Args:
            teacher_name: 教师姓名
            school_code: 学校编码
            classroom_code: 班级编码（可选）
            is_duplicate: 是否为重名（重名时需要添加学校代码和班级编码）

        Returns:
            用户名字符串
        """
        name_pinyin = TeacherAssignmentImportService._name_to_pinyin(teacher_name)
        
        if is_duplicate and classroom_code:
            # 重名时使用：姓名拼音_学校代码_班级编码
            username = f"{name_pinyin}_{school_code}_{classroom_code}"
        else:
            # 非重名时直接使用姓名拼音
            username = name_pinyin
        
        # 清理特殊字符，只保留字母、数字、下划线
        username = re.sub(r"[^a-z0-9_]", "", username)
        
        return username

    @staticmethod
    async def _ensure_unique_username(
        db: AsyncSession,
        base_username: str,
    ) -> str:
        """
        确保用户名唯一，如果已存在则添加序号

        Args:
            db: 数据库会话
            base_username: 基础用户名

        Returns:
            唯一的用户名
        """
        username = base_username
        counter = 1
        
        while True:
            result = await db.execute(select(User).where(User.username == username))
            if not result.scalar_one_or_none():
                return username
            
            # 如果已存在，添加序号
            username = f"{base_username}_{counter}"
            counter += 1

    @staticmethod
    async def _ensure_unique_email(
        db: AsyncSession,
        base_email: str,
    ) -> str:
        """
        确保邮箱唯一，如果已存在则添加序号

        Args:
            db: 数据库会话
            base_email: 基础邮箱

        Returns:
            唯一的邮箱
        """
        email = base_email
        counter = 1
        
        while True:
            result = await db.execute(select(User).where(User.email == email))
            if not result.scalar_one_or_none():
                return email
            
            # 如果已存在，添加序号
            email_local, email_domain = base_email.split("@")
            email = f"{email_local}_{counter}@{email_domain}"
            counter += 1

    @staticmethod
    async def _check_duplicate_teacher_name(
        db: AsyncSession,
        teacher_name: str,
        school_id: int,
        grade_id: int,
        classroom_code: Optional[str],
    ) -> bool:
        """
        检查是否存在同名教师（在同一学校、年级、班级）

        Args:
            db: 数据库会话
            teacher_name: 教师姓名
            school_id: 学校ID
            grade_id: 年级ID
            classroom_code: 班级编码

        Returns:
            如果存在同名教师返回True，否则返回False
        """
        # 查找同一学校、年级、班级的同名教师
        query = select(User).where(
            and_(
                User.role == UserRole.TEACHER,
                User.full_name == teacher_name,
                User.school_id == school_id,
                User.grade_id == grade_id,
            )
        )
        
        # 如果有班级编码，进一步筛选（通过查找该班级的教师）
        if classroom_code:
            from app.models import Classroom
            # 查找该班级
            classroom_result = await db.execute(
                select(Classroom).where(
                    and_(
                        Classroom.school_id == school_id,
                        Classroom.grade_id == grade_id,
                        Classroom.code == classroom_code,
                    )
                )
            )
            classroom = classroom_result.scalar_one_or_none()
            if classroom:
                # 检查该班级的教师
                query = query.where(User.classroom_id == classroom.id)
        
        result = await db.execute(query)
        existing_teacher = result.scalar_one_or_none()
        
        return existing_teacher is not None

    @staticmethod
    async def create_teacher_if_not_exists(
        db: AsyncSession,
        teacher_name: str,
        school: School,
        grade: Grade,
        classroom: Optional[Classroom] = None,
        default_password: str = "Teacher@123456",
    ) -> Tuple[User, bool]:
        """
        如果教师不存在则创建，返回教师对象和是否为新创建

        Args:
            db: 数据库会话
            teacher_name: 教师姓名
            school: 学校对象
            grade: 年级对象
            classroom: 班级对象（可选）
            default_password: 默认密码

        Returns:
            (教师对象, 是否为新创建)
        """
        # 先尝试查找教师
        teacher = await TeacherAssignmentImportService.find_teacher(db, teacher_name)
        if teacher:
            return teacher, False
        
        # 检查是否存在同名教师（用于判断是否需要添加学校代码和班级编码）
        classroom_code_str: Optional[str] = None
        if classroom and classroom.code:  # type: ignore
            classroom_code_str = str(classroom.code)  # type: ignore
        is_duplicate = await TeacherAssignmentImportService._check_duplicate_teacher_name(
            db, teacher_name, int(school.id), int(grade.id), classroom_code_str  # type: ignore
        )
        
        # 生成用户名
        base_username = TeacherAssignmentImportService._generate_username(
            teacher_name=teacher_name,
            school_code=str(school.code),
            classroom_code=classroom_code_str,
            is_duplicate=is_duplicate,
        )
        username = await TeacherAssignmentImportService._ensure_unique_username(
            db, base_username
        )
        
        # 生成邮箱
        base_email = f"{username}@inspireed.local"
        email = await TeacherAssignmentImportService._ensure_unique_email(db, base_email)
        
        # 创建教师
        teacher = User(
            username=username,
            email=email,
            hashed_password=get_password_hash(default_password),
            full_name=teacher_name,
            role=UserRole.TEACHER,
            is_active=True,
            school_id=int(school.id),  # type: ignore
            region_id=int(school.region_id) if school.region_id else None,  # type: ignore
            grade_id=int(grade.id) if grade else None,  # type: ignore
            classroom_id=int(classroom.id) if classroom else None,  # type: ignore
        )
        
        db.add(teacher)
        await db.flush()  # 刷新以获取ID
        
        logger.info(
            f"自动创建教师：{teacher_name} (username={username}, email={email})"
        )
        
        return teacher, True

    @staticmethod
    async def find_school(
        db: AsyncSession,
        school_name: str,
    ) -> Optional[School]:
        """
        根据学校名称查找学校

        Args:
            db: 数据库会话
            school_name: 学校名称

        Returns:
            学校对象，如果未找到返回None
        """
        if not school_name:
            return None

        school_name = str(school_name).strip()

        result = await db.execute(
            select(School).where(
                or_(
                    School.name == school_name,
                    School.name.ilike(f"%{school_name}%"),
                )
            )
        )
        school = result.scalar_one_or_none()

        return school

    @staticmethod
    async def find_grade(
        db: AsyncSession,
        grade_name: Optional[str] = None,
        grade_level: Optional[int] = None,
    ) -> Optional[Grade]:
        """
        根据年级名称或年级级别查找年级

        Args:
            db: 数据库会话
            grade_name: 年级名称（如"七年级"）
            grade_level: 年级级别（如7）

        Returns:
            年级对象，如果未找到返回None
        """
        # 优先使用年级级别查找
        if grade_level is not None:
            try:
                grade_level_int = int(grade_level)
                result = await db.execute(
                    select(Grade).where(Grade.level == grade_level_int)
                )
                grade = result.scalar_one_or_none()
                if grade:
                    return grade
            except (ValueError, TypeError):
                pass
        
        # 如果年级级别查找失败，尝试通过名称查找
        if grade_name:
            grade_name = str(grade_name).strip()
            result = await db.execute(
                select(Grade).where(
                    or_(
                        Grade.name == grade_name,
                        Grade.name.ilike(f"%{grade_name}%"),
                    )
                )
            )
            grade = result.scalar_one_or_none()
            return grade

        return None

    @staticmethod
    async def find_classroom(
        db: AsyncSession,
        school_id: int,
        grade_id: int,
        classroom_name: Optional[str] = None,
        classroom_code: Optional[str] = None,
    ) -> Optional[Classroom]:
        """
        根据班级名称或班级编码查找班级

        Args:
            db: 数据库会话
            school_id: 学校ID
            grade_id: 年级ID
            classroom_name: 班级名称（如"七年级1班"）
            classroom_code: 班级编码（如"701"）

        Returns:
            班级对象，如果未找到返回None
        """
        # 优先使用班级编码查找
        if classroom_code:
            classroom_code = str(classroom_code).strip()
            result = await db.execute(
                select(Classroom).where(
                    and_(
                        Classroom.school_id == school_id,
                        Classroom.grade_id == grade_id,
                        Classroom.code == classroom_code,
                    )
                )
            )
            classroom = result.scalar_one_or_none()
            if classroom:
                return classroom
        
        # 如果班级编码查找失败，尝试通过名称查找
        if classroom_name:
            classroom_name = str(classroom_name).strip()
            result = await db.execute(
                select(Classroom).where(
                    and_(
                        Classroom.school_id == school_id,
                        Classroom.grade_id == grade_id,
                        or_(
                            Classroom.name == classroom_name,
                            Classroom.name.ilike(f"%{classroom_name}%"),
                        ),
                    )
                )
            )
            classroom = result.scalar_one_or_none()
            return classroom

        return None

    @staticmethod
    async def find_subject(
        db: AsyncSession,
        subject_name: str,
    ) -> Optional[Subject]:
        """
        根据学科名称查找学科

        Args:
            db: 数据库会话
            subject_name: 学科名称

        Returns:
            学科对象，如果未找到返回None
        """
        if not subject_name:
            return None

        subject_name = str(subject_name).strip()

        result = await db.execute(
            select(Subject).where(
                or_(
                    Subject.name == subject_name,
                    Subject.name.ilike(f"%{subject_name}%"),
                )
            )
        )
        subject = result.scalar_one_or_none()

        return subject

    @staticmethod
    async def find_semester(
        db: AsyncSession,
        semester_name: Optional[str] = None,
        semester_number: Optional[int] = None,
        academic_year: Optional[str] = None,
    ) -> Optional[Semester]:
        """
        根据学期名称或学期编号查找学期

        Args:
            db: 数据库会话
            semester_name: 学期名称（如"2024-2025学年第一学期"）
            semester_number: 学期编号（1或2，1表示上学期/up，2表示下学期/down）
            academic_year: 学年（如"2024-2025"），用于学期编号查找

        Returns:
            学期对象，如果未找到返回None
        """
        # 优先使用学期编号查找（需要学年）
        if semester_number is not None and academic_year:
            try:
                semester_number_int = int(semester_number)
                if semester_number_int == 1:
                    semester_type = "up"
                elif semester_number_int == 2:
                    semester_type = "down"
                else:
                    semester_type = None
                
                if semester_type:
                    result = await db.execute(
                        select(Semester).where(
                            and_(
                                Semester.year == academic_year,
                                Semester.semester_type == semester_type,
                            )
                        )
                    )
                    semester = result.scalar_one_or_none()
                    if semester:
                        return semester
            except (ValueError, TypeError):
                pass
        
        # 如果学期编号查找失败，尝试通过名称查找
        if semester_name:
            semester_name = str(semester_name).strip()
            result = await db.execute(
                select(Semester).where(
                    or_(
                        Semester.name == semester_name,
                        Semester.name.ilike(f"%{semester_name}%"),
                    )
                )
            )
            semester = result.scalar_one_or_none()
            return semester

        return None

    @staticmethod
    def _parse_academic_year(academic_year: str) -> Tuple[int, int]:
        """
        解析学年字符串，返回开始年份和结束年份

        Args:
            academic_year: 学年字符串，如 "2025-2026"

        Returns:
            (开始年份, 结束年份)

        Raises:
            ValueError: 如果学年格式不正确
        """
        try:
            parts = academic_year.split("-")
            if len(parts) != 2:
                raise ValueError(f"学年格式错误：{academic_year}，应为 YYYY-YYYY 格式")
            start_year = int(parts[0])
            end_year = int(parts[1])
            if end_year != start_year + 1:
                raise ValueError(f"学年格式错误：{academic_year}，结束年份应为开始年份+1")
            return start_year, end_year
        except (ValueError, IndexError) as e:
            raise ValueError(f"学年格式错误：{academic_year}，应为 YYYY-YYYY 格式") from e

    @staticmethod
    def _calculate_semester_dates(
        start_year: int,
        semester_type: str,
    ) -> Tuple[datetime, datetime]:
        """
        计算学期的开始和结束日期

        Args:
            start_year: 学年开始年份
            semester_type: 学期类型，"up" 或 "down"

        Returns:
            (开始日期, 结束日期)
        """
        if semester_type == "up":
            # 上学期：9月1日 到 次年1月31日
            start_date = datetime(start_year, 9, 1)
            end_date = datetime(start_year + 1, 1, 31)
        elif semester_type == "down":
            # 下学期：2月1日 到 7月31日
            start_date = datetime(start_year + 1, 2, 1)
            end_date = datetime(start_year + 1, 7, 31)
        else:
            raise ValueError(f"无效的学期类型：{semester_type}，应为 'up' 或 'down'")
        
        return start_date, end_date

    @staticmethod
    def _generate_semester_name(academic_year: str, semester_type: str) -> str:
        """
        生成学期名称

        Args:
            academic_year: 学年字符串，如 "2025-2026"
            semester_type: 学期类型，"up" 或 "down"

        Returns:
            学期名称，如 "2025-2026学年上学期"
        """
        if semester_type == "up":
            return f"{academic_year}学年上学期"
        elif semester_type == "down":
            return f"{academic_year}学年下学期"
        else:
            raise ValueError(f"无效的学期类型：{semester_type}")

    @staticmethod
    async def create_semester_if_not_exists(
        db: AsyncSession,
        academic_year: str,
        semester_number: int,
        region_id: Optional[int] = None,
    ) -> Tuple[Semester, bool]:
        """
        如果学期不存在则创建，返回学期对象和是否为新创建

        Args:
            db: 数据库会话
            academic_year: 学年字符串，如 "2025-2026"
            semester_number: 学期编号（1表示上学期，2表示下学期）
            region_id: 区县ID（可选）

        Returns:
            (学期对象, 是否为新创建)
        """
        # 转换学期编号
        if semester_number == 1:
            semester_type = "up"
        elif semester_number == 2:
            semester_type = "down"
        else:
            raise ValueError(f"无效的学期编号：{semester_number}，应为 1 或 2")

        # 先尝试查找学期
        semester = await TeacherAssignmentImportService.find_semester(
            db,
            semester_number=semester_number,
            academic_year=academic_year,
        )
        if semester:
            return semester, False

        # 解析学年并计算日期
        start_year, end_year = TeacherAssignmentImportService._parse_academic_year(academic_year)
        start_date, end_date = TeacherAssignmentImportService._calculate_semester_dates(
            start_year, semester_type
        )
        semester_name = TeacherAssignmentImportService._generate_semester_name(
            academic_year, semester_type
        )

        # 创建学期
        semester = Semester(
            year=academic_year,
            semester_type=semester_type,
            name=semester_name,
            start_date=start_date,
            end_date=end_date,
            is_current=False,  # 默认不是当前学期
            region_id=region_id,
            is_active=True,
        )

        db.add(semester)
        await db.flush()  # 刷新以获取ID

        logger.info(
            f"自动创建学期：{semester_name} (year={academic_year}, type={semester_type})"
        )

        return semester, True

    @staticmethod
    async def find_position_type(
        db: AsyncSession,
        position_name: str,
    ) -> Optional[TeacherPositionType]:
        """
        根据职务名称查找职务类型

        Args:
            db: 数据库会话
            position_name: 职务名称（如"班主任"、"学科教师"、"校长"等）

        Returns:
            职务类型对象，如果未找到返回None
        """
        if not position_name:
            return None

        position_name = str(position_name).strip()

        # 先尝试精确匹配
        result = await db.execute(
            select(TeacherPositionType).where(
                and_(
                    TeacherPositionType.name == position_name,
                    TeacherPositionType.is_active == True,
                )
            )
        )
        position_type = result.scalar_one_or_none()
        if position_type:
            return position_type

        # 尝试模糊匹配
        result = await db.execute(
            select(TeacherPositionType).where(
                and_(
                    TeacherPositionType.name.ilike(f"%{position_name}%"),
                    TeacherPositionType.is_active == True,
                )
            )
        )
        position_type = result.scalar_one_or_none()
        if position_type:
            return position_type

        # 尝试通过代码匹配（支持旧的枚举值）
        position_name_lower = position_name.lower()
        if position_name_lower in ["班主任", "head_teacher", "headteacher"]:
            result = await db.execute(
                select(TeacherPositionType).where(
                    and_(
                        TeacherPositionType.code == "head_teacher",
                        TeacherPositionType.is_active == True,
                    )
                )
            )
            position_type = result.scalar_one_or_none()
            if position_type:
                return position_type
        elif position_name_lower in ["学科教师", "subject_teacher", "subjectteacher", "任课教师"]:
            result = await db.execute(
                select(TeacherPositionType).where(
                    and_(
                        TeacherPositionType.code == "subject_teacher",
                        TeacherPositionType.is_active == True,
                    )
                )
            )
            position_type = result.scalar_one_or_none()
            if position_type:
                return position_type

        return None

    @staticmethod
    def parse_assignment_type(assignment_type_str: str) -> Optional[TeachingAssignmentType]:
        """
        解析任务类型字符串

        Args:
            assignment_type_str: 任务类型字符串

        Returns:
            任务类型枚举值，如果无法解析返回None
        """
        if not assignment_type_str:
            return None

        assignment_type_str = str(assignment_type_str).strip().lower()

        # 支持多种格式
        if assignment_type_str in ["班主任", "head_teacher", "headteacher"]:
            return TeachingAssignmentType.HEAD_TEACHER
        elif assignment_type_str in ["学科教师", "subject_teacher", "subjectteacher", "任课教师"]:
            return TeachingAssignmentType.SUBJECT_TEACHER

        return None

    @staticmethod
    def parse_is_active(is_active_str: Optional[str]) -> bool:
        """
        解析是否激活字段

        Args:
            is_active_str: 是否激活字符串

        Returns:
            True表示激活，False表示非激活
        """
        if not is_active_str:
            return True  # 默认激活

        is_active_str = str(is_active_str).strip().lower()

        # 支持多种格式
        if is_active_str in ["是", "yes", "true", "1", "激活", "启用"]:
            return True
        elif is_active_str in ["否", "no", "false", "0", "非激活", "停用"]:
            return False

        return True  # 默认激活

    @staticmethod
    async def import_assignments(
        db: AsyncSession,
        file_path: Path,
        update_existing: bool = False,
        auto_create_teachers: bool = False,
        auto_create_semesters: bool = False,
    ) -> Dict[str, Any]:
        """
        批量导入教师教学任务

        Args:
            db: 数据库会话
            file_path: Excel文件路径
            update_existing: 是否更新已存在的任务
            auto_create_teachers: 如果教师不存在，是否自动创建
            auto_create_semesters: 如果学期不存在，是否自动创建

        Returns:
            导入结果字典，包含：
            - total: 总记录数
            - success: 成功数
            - failed: 失败数
            - created: 创建数
            - updated: 更新数
            - skipped: 跳过数
            - errors: 错误列表
            - created_teachers: 新创建的教师列表（如果启用自动创建）
            - created_semesters: 新创建的学期列表（如果启用自动创建）
        """
        result = {
            "total": 0,
            "success": 0,
            "failed": 0,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [],
            "created_teachers": [],  # 新创建的教师列表
            "created_semesters": [],  # 新创建的学期列表
        }
        
        # 用于跟踪本批次已创建的教师和学期，避免重复创建
        created_teachers_cache: Dict[str, User] = {}
        created_semesters_cache: Dict[str, Semester] = {}

        try:
            # 解析Excel
            data = TeacherAssignmentImportService.parse_assignment_excel(file_path)
            result["total"] = len(data)

            for row_data in data:
                row_number = row_data.get("_row_number", 0)

                try:
                    # 读取字段（支持新旧格式）
                    teacher_name = row_data.get("teacher_name")
                    school_name = row_data.get("school_name")
                    grade_name = row_data.get("grade_name")
                    grade_level = row_data.get("grade_level")
                    classroom_name = row_data.get("classroom_name")
                    classroom_code = row_data.get("classroom_code")
                    subject_name = row_data.get("subject_name")
                    semester_name = row_data.get("semester_name")
                    semester_number = row_data.get("semester_number")
                    academic_year = row_data.get("academic_year")
                    assignment_type_str = row_data.get("assignment_type")

                    # 验证必填字段
                    if not teacher_name or not school_name or not subject_name or not academic_year or not assignment_type_str:
                        result["errors"].append({
                            "row": row_number,
                            "field": None,
                            "message": "缺少必填字段（教师姓名、学校名称、学科名称、学年、任务类型）",
                        })
                        result["failed"] += 1
                        continue

                    # 验证年级字段（至少有一个）
                    if not grade_name and grade_level is None:
                        result["errors"].append({
                            "row": row_number,
                            "field": "年级",
                            "message": "缺少年级字段（需要：年级名称 或 年级级别）",
                        })
                        result["failed"] += 1
                        continue

                    # 验证班级字段（至少有一个）
                    if not classroom_name and not classroom_code:
                        result["errors"].append({
                            "row": row_number,
                            "field": "班级",
                            "message": "缺少班级字段（需要：班级名称 或 班级编码）",
                        })
                        result["failed"] += 1
                        continue

                    # 验证学期字段（至少有一个）
                    if not semester_name and semester_number is None:
                        result["errors"].append({
                            "row": row_number,
                            "field": "学期",
                            "message": "缺少学期字段（需要：学期名称 或 学期编号）",
                        })
                        result["failed"] += 1
                        continue

                    # 先查找学校（创建教师需要学校信息）
                    school = await TeacherAssignmentImportService.find_school(db, school_name)
                    if not school:
                        result["errors"].append({
                            "row": row_number,
                            "field": "学校名称",
                            "message": f"未找到学校：{school_name}",
                        })
                        result["failed"] += 1
                        continue

                    # 解析年级级别（如果提供的是数字字符串）
                    grade_level_int = None
                    if grade_level is not None:
                        try:
                            grade_level_int = int(grade_level)
                        except (ValueError, TypeError):
                            pass

                    # 查找年级（创建教师需要年级信息）
                    grade = await TeacherAssignmentImportService.find_grade(
                        db, grade_name=grade_name, grade_level=grade_level_int
                    )
                    if not grade:
                        grade_display = grade_level_int if grade_level_int is not None else grade_name
                        result["errors"].append({
                            "row": row_number,
                            "field": "年级",
                            "message": f"未找到年级：{grade_display}",
                        })
                        result["failed"] += 1
                        continue

                    # 查找班级（创建教师可能需要班级信息）
                    classroom = await TeacherAssignmentImportService.find_classroom(
                        db,
                        school_id=int(school.id),  # type: ignore
                        grade_id=int(grade.id),  # type: ignore
                        classroom_name=classroom_name,
                        classroom_code=classroom_code,
                    )
                    if not classroom:
                        classroom_display = classroom_code if classroom_code else classroom_name
                        result["errors"].append({
                            "row": row_number,
                            "field": "班级",
                            "message": f"未找到班级：{classroom_display}（学校：{school_name}，年级：{grade.name}）",
                        })
                        result["failed"] += 1
                        continue

                    # 查找或创建教师
                    teacher_key = f"{teacher_name}_{school.id}_{grade.id}_{classroom_code or ''}"
                    
                    # 先检查缓存
                    if teacher_key in created_teachers_cache:
                        teacher = created_teachers_cache[teacher_key]
                    else:
                        teacher = await TeacherAssignmentImportService.find_teacher(db, teacher_name)
                        
                        if not teacher:
                            if auto_create_teachers:
                                # 自动创建教师
                                try:
                                    teacher, is_new = await TeacherAssignmentImportService.create_teacher_if_not_exists(
                                        db=db,
                                        teacher_name=teacher_name,
                                        school=school,
                                        grade=grade,
                                        classroom=classroom,
                                    )
                                    
                                    if is_new:
                                        # 记录新创建的教师信息
                                        # 确保 classroom_code 是字符串类型
                                        classroom_code_str: Optional[str] = None
                                        if classroom_code is not None:
                                            classroom_code_str = str(classroom_code)
                                        
                                        result["created_teachers"].append({
                                            "teacher_name": teacher_name,
                                            "username": teacher.username,
                                            "email": teacher.email,
                                            "password": "Teacher@123456",  # 默认密码
                                            "school_name": school_name,
                                            "school_code": str(school.code),  # type: ignore
                                            "grade_name": grade.name if grade.name else None,  # type: ignore
                                            "classroom_code": classroom_code_str,
                                            "classroom_name": classroom.name if classroom else None,  # type: ignore
                                            "row_number": row_number,
                                        })
                                        created_teachers_cache[teacher_key] = teacher
                                except Exception as e:
                                    logger.error(f"自动创建教师失败：{str(e)}", exc_info=True)
                                    result["errors"].append({
                                        "row": row_number,
                                        "field": "教师姓名",
                                        "message": f"自动创建教师失败：{str(e)}",
                                    })
                                    result["failed"] += 1
                                    continue
                            else:
                                result["errors"].append({
                                    "row": row_number,
                                    "field": "教师姓名",
                                    "message": f"未找到教师：{teacher_name}",
                                })
                                result["failed"] += 1
                                continue
                        else:
                            # 缓存已存在的教师
                            created_teachers_cache[teacher_key] = teacher

                    # 查找学科
                    subject = await TeacherAssignmentImportService.find_subject(db, subject_name)
                    if not subject:
                        result["errors"].append({
                            "row": row_number,
                            "field": "学科名称",
                            "message": f"未找到学科：{subject_name}",
                        })
                        result["failed"] += 1
                        continue

                    # 解析学期编号（如果提供的是数字字符串）
                    semester_number_int = None
                    if semester_number is not None:
                        try:
                            semester_number_int = int(semester_number)
                        except (ValueError, TypeError):
                            pass

                    # 查找或创建学期
                    semester_key = f"{academic_year}_{semester_number_int}"
                    
                    # 先检查缓存
                    if semester_key in created_semesters_cache:
                        semester = created_semesters_cache[semester_key]
                    else:
                        semester = await TeacherAssignmentImportService.find_semester(
                            db,
                            semester_name=semester_name,
                            semester_number=semester_number_int,
                            academic_year=academic_year,
                        )
                        
                        if not semester:
                            # 记录调试信息
                            logger.debug(
                                f"学期未找到 - auto_create_semesters={auto_create_semesters}, "
                                f"semester_number_int={semester_number_int}, academic_year={academic_year}"
                            )
                            
                            if auto_create_semesters and semester_number_int is not None and academic_year:
                                # 自动创建学期
                                logger.info(
                                    f"开始自动创建学期 - 学年={academic_year}, 学期编号={semester_number_int}"
                                )
                                try:
                                    # 获取学校所属区域（如果有）
                                    region_id = int(school.region_id) if school.region_id else None  # type: ignore
                                    
                                    semester, is_new = await TeacherAssignmentImportService.create_semester_if_not_exists(
                                        db=db,
                                        academic_year=academic_year,
                                        semester_number=semester_number_int,
                                        region_id=region_id,
                                    )
                                    
                                    # 无论是否为新创建，都缓存学期对象
                                    created_semesters_cache[semester_key] = semester
                                    
                                    if is_new:
                                        # 记录新创建的学期信息
                                        result["created_semesters"].append({
                                            "semester_name": semester.name,  # type: ignore
                                            "academic_year": academic_year,
                                            "semester_number": semester_number_int,
                                            "semester_type": semester.semester_type,  # type: ignore
                                            "start_date": semester.start_date.isoformat() if semester.start_date else None,  # type: ignore
                                            "end_date": semester.end_date.isoformat() if semester.end_date else None,  # type: ignore
                                            "row_number": row_number,
                                        })
                                except Exception as e:
                                    logger.error(f"自动创建学期失败：{str(e)}", exc_info=True)
                                    semester_display = semester_number_int if semester_number_int is not None else semester_name
                                    result["errors"].append({
                                        "row": row_number,
                                        "field": "学期",
                                        "message": f"自动创建学期失败：{str(e)}",
                                    })
                                    result["failed"] += 1
                                    continue
                            else:
                                semester_display = semester_number_int if semester_number_int is not None else semester_name
                                result["errors"].append({
                                    "row": row_number,
                                    "field": "学期",
                                    "message": f"未找到学期：{semester_display}（学年：{academic_year}）",
                                })
                                result["failed"] += 1
                                continue
                        else:
                            # 缓存已存在的学期
                            created_semesters_cache[semester_key] = semester

                    # 查找职务类型（优先使用新的职务类型系统）
                    position_type = await TeacherAssignmentImportService.find_position_type(
                        db, assignment_type_str
                    )
                    if not position_type:
                        result["errors"].append({
                            "row": row_number,
                            "field": "任务类型",
                            "message": f"未找到职务类型：{assignment_type_str}（请先在职务类型管理中创建该职务）",
                        })
                        result["failed"] += 1
                        continue
                    
                    # 向后兼容：如果找不到职务类型，尝试使用旧的枚举（已废弃）
                    assignment_type = None
                    if not position_type:
                        assignment_type = TeacherAssignmentImportService.parse_assignment_type(assignment_type_str)
                        if not assignment_type:
                            result["errors"].append({
                                "row": row_number,
                                "field": "任务类型",
                                "message": f"未找到职务类型：{assignment_type_str}（请先在职务类型管理中创建该职务）",
                            })
                            result["failed"] += 1
                            continue

                    # 解析是否激活
                    is_active = TeacherAssignmentImportService.parse_is_active(row_data.get("is_active"))

                    # 检查是否已存在
                    existing = await db.execute(
                        select(TeacherTeachingAssignment).where(
                            and_(
                                TeacherTeachingAssignment.teacher_id == teacher.id,
                                TeacherTeachingAssignment.semester_id == semester.id,
                                TeacherTeachingAssignment.classroom_id == classroom.id,
                                TeacherTeachingAssignment.subject_id == subject.id,
                            )
                        )
                    )
                    existing_assignment = existing.scalar_one_or_none()

                    if existing_assignment:
                        if update_existing:
                            # 更新现有任务
                            existing_assignment.school_id = int(school.id)  # type: ignore
                            existing_assignment.grade_id = int(grade.id)  # type: ignore
                            existing_assignment.academic_year = str(academic_year)  # type: ignore
                            # 优先使用新的职务类型
                            if position_type:
                                existing_assignment.position_type_id = int(position_type.id)  # type: ignore
                                existing_assignment.assignment_type = None  # type: ignore
                            else:
                                existing_assignment.assignment_type = assignment_type  # type: ignore
                            existing_assignment.is_active = bool(is_active)  # type: ignore
                            db.add(existing_assignment)
                            result["updated"] += 1
                            result["success"] += 1
                        else:
                            # 跳过已存在的任务
                            result["skipped"] += 1
                            result["success"] += 1
                    else:
                        # 创建新任务
                        new_assignment = TeacherTeachingAssignment(
                            teacher_id=int(teacher.id),  # type: ignore
                            school_id=int(school.id),  # type: ignore
                            grade_id=int(grade.id),  # type: ignore
                            classroom_id=int(classroom.id),  # type: ignore
                            subject_id=int(subject.id),  # type: ignore
                            semester_id=int(semester.id),  # type: ignore
                            academic_year=str(academic_year),
                            position_type_id=int(position_type.id) if position_type else None,  # type: ignore
                            assignment_type=assignment_type,  # 向后兼容
                            is_active=bool(is_active),
                        )
                        db.add(new_assignment)
                        result["created"] += 1
                        result["success"] += 1

                except Exception as e:
                    logger.error(f"导入第 {row_number} 行数据时出错：{str(e)}", exc_info=True)
                    result["errors"].append({
                        "row": row_number,
                        "field": None,
                        "message": f"处理失败：{str(e)}",
                    })
                    result["failed"] += 1

            # 提交事务
            await db.commit()

        except TeacherAssignmentImportServiceError as e:
            await db.rollback()
            result["errors"].append({
                "row": 0,
                "field": None,
                "message": str(e),
            })
            result["failed"] = result["total"]
        except Exception as e:
            await db.rollback()
            logger.error(f"批量导入教师教学任务失败：{str(e)}", exc_info=True)
            result["errors"].append({
                "row": 0,
                "field": None,
                "message": f"导入失败：{str(e)}",
            })
            result["failed"] = result["total"]

        return result
