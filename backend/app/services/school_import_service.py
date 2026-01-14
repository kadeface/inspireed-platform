"""
学校批量导入服务

提供Excel文件解析、区域/学校匹配和批量导入功能
"""

import logging
import hashlib
from typing import Any, Dict, List, Optional, Tuple
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_
from sqlalchemy.exc import IntegrityError

from app.models import Region, School

logger = logging.getLogger(__name__)


class SchoolImportServiceError(Exception):
    """学校导入服务错误"""
    pass


class SchoolImportService:
    """学校批量导入服务"""

    # Excel列名映射（支持多种格式）
    # 注意：列名匹配时会自动去除星号等标记，所以这里不需要包含带星号的版本
    COLUMN_MAPPING = {
        # 区域名称（支持：区域名称、区域名称*、市(区)、区域、市、区）
        "区域名称": "region_name",
        "市(区)": "region_name",
        "区域": "region_name",
        "市": "region_name",
        "区": "region_name",
        # 学校名称（支持：学校名称、学校名称*、学校）
        "学校名称": "school_name",
        "学校": "school_name",
        # 学校代码
        "学校代码": "school_code",
        "代码": "school_code",
        # 学校类型
        "学校类型": "school_type",
        "类型": "school_type",
        # 地址
        "地址": "address",
        "学校地址": "address",
        # 联系电话
        "联系电话": "phone",
        "电话": "phone",
        "联系方式": "phone",
        # 邮箱
        "邮箱": "email",
        "Email": "email",
        "email": "email",
        # 校长
        "校长": "principal",
        "校长姓名": "principal",
    }

    # 必需列
    REQUIRED_COLUMNS = ["区域名称", "学校名称"]

    @staticmethod
    async def parse_school_excel(file_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        解析Excel文件

        Args:
            file_path: Excel文件路径

        Returns:
            (记录列表, 错误列表)
        """
        records = []
        errors = []

        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise SchoolImportServiceError("Excel文件为空或没有工作表")

            # 获取表头
            headers = [cell.value for cell in ws[1]]  # type: ignore
            if not headers:
                raise SchoolImportServiceError("Excel文件为空或没有表头")

            # 标准化列名（去除空格和星号等标记）
            headers = [str(h).strip().rstrip('*').strip() if h else "" for h in headers]

            # 查找列索引
            column_indices = {}
            for col_name, field_name in SchoolImportService.COLUMN_MAPPING.items():
                if col_name in headers:
                    column_indices[field_name] = headers.index(col_name)

            # 验证必需列
            required_fields = ["region_name", "school_name"]
            missing_fields = [f for f in required_fields if f not in column_indices]
            if missing_fields:
                raise SchoolImportServiceError(f"缺少必需列: {missing_fields}")

            # 解析数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # type: ignore
                if not any(row):  # 跳过空行
                    continue

                try:
                    record: Dict[str, Any] = {
                        "row_number": row_idx,
                    }

                    # 提取区域名称（必需）
                    region_name = row[column_indices["region_name"]]  # type: ignore
                    if not region_name:
                        errors.append({
                            "row": row_idx,
                            "field": "区域名称",
                            "message": "区域名称不能为空"
                        })
                        continue
                    record["region_name"] = str(region_name).strip()

                    # 提取学校名称（必需）
                    school_name = row[column_indices["school_name"]]  # type: ignore
                    if not school_name:
                        errors.append({
                            "row": row_idx,
                            "field": "学校名称",
                            "message": "学校名称不能为空"
                        })
                        continue
                    record["school_name"] = str(school_name).strip()

                    # 提取可选字段
                    if "school_code" in column_indices:
                        school_code = row[column_indices["school_code"]]  # type: ignore
                        record["school_code"] = str(school_code).strip() if school_code else None
                    else:
                        record["school_code"] = None

                    if "school_type" in column_indices:
                        school_type = row[column_indices["school_type"]]  # type: ignore
                        record["school_type"] = str(school_type).strip() if school_type else None
                    else:
                        record["school_type"] = None

                    if "address" in column_indices:
                        address = row[column_indices["address"]]  # type: ignore
                        record["address"] = str(address).strip() if address else None
                    else:
                        record["address"] = None

                    if "phone" in column_indices:
                        phone = row[column_indices["phone"]]  # type: ignore
                        record["phone"] = str(phone).strip() if phone else None
                    else:
                        record["phone"] = None

                    if "email" in column_indices:
                        email = row[column_indices["email"]]  # type: ignore
                        record["email"] = str(email).strip() if email else None
                    else:
                        record["email"] = None

                    if "principal" in column_indices:
                        principal = row[column_indices["principal"]]  # type: ignore
                        record["principal"] = str(principal).strip() if principal else None
                    else:
                        record["principal"] = None

                    records.append(record)

                except Exception as e:
                    errors.append({
                        "row": row_idx,
                        "field": None,
                        "message": f"解析行数据失败: {str(e)}"
                    })

        except SchoolImportServiceError:
            raise
        except Exception as e:
            raise SchoolImportServiceError(f"解析Excel文件失败: {str(e)}")

        return records, errors

    @staticmethod
    async def find_or_create_region(
        db: AsyncSession,
        region_name: str,
        auto_create: bool = True
    ) -> Optional[Region]:
        """
        查找或创建区域

        Args:
            db: 数据库会话
            region_name: 区域名称
            auto_create: 是否自动创建（如果不存在）

        Returns:
            Region对象，如果不存在且auto_create=False则返回None
        """
        # 1. 精确匹配
        result = await db.execute(
            select(Region).where(Region.name == region_name)
        )
        region = result.scalar_one_or_none()
        if region:
            return region

        # 2. 模糊匹配
        result = await db.execute(
            select(Region).where(Region.name.ilike(f"%{region_name}%"))
        )
        region = result.scalar_one_or_none()
        if region:
            return region

        # 3. 自动创建
        if auto_create:
            # 生成唯一code
            region_code = await SchoolImportService._generate_region_code(region_name, db)
            
            # 推断level
            level = SchoolImportService._infer_region_level(region_name)
            
            region = Region(
                name=region_name,
                code=region_code,
                level=level,
                is_active=True
            )
            db.add(region)
            await db.flush()
            return region

        return None

    @staticmethod
    async def find_or_create_school(
        db: AsyncSession,
        school_data: Dict[str, Any],
        region_id: int
    ) -> Tuple[Optional[School], str]:
        """
        查找或创建学校

        Args:
            db: 数据库会话
            school_data: 学校数据字典
            region_id: 区域ID
            auto_create: 是否自动创建（如果不存在）

        Returns:
            (School对象, 操作类型: 'created'/'updated'/'skipped'/'error')
        """
        school_name = school_data["school_name"]
        school_code = school_data.get("school_code")

        # 1. 按学校代码精确匹配
        if school_code:
            result = await db.execute(
                select(School).where(School.code == school_code)
            )
            school = result.scalar_one_or_none()
            if school:
                # 更新可选字段
                updated = False
                for field in ["school_type", "address", "phone", "email", "principal"]:
                    if field in school_data and school_data[field]:
                        setattr(school, field, school_data[field])
                        updated = True
                if updated:
                    await db.flush()
                    return school, "updated"
                return school, "skipped"

        # 2. 按名称+区域匹配
        result = await db.execute(
            select(School).where(
                School.name == school_name,
                School.region_id == region_id
            )
        )
        school = result.scalar_one_or_none()
        if school:
            # 更新可选字段
            updated = False
            for field in ["school_type", "address", "phone", "email", "principal"]:
                if field in school_data and school_data[field]:
                    setattr(school, field, school_data[field])
                    updated = True
            if updated:
                await db.flush()
                return school, "updated"
            return school, "skipped"

        # 3. 创建新学校
        # 生成学校代码（如果没有提供）
        if not school_code:
            school_code = await SchoolImportService._generate_school_code(
                school_name, region_id, db
            )

        # 推断学校类型（如果没有提供）
        school_type = school_data.get("school_type")
        if not school_type:
            school_type = "高中"  # 默认值

        school = School(
            name=school_name,
            code=school_code,
            region_id=region_id,
            school_type=school_type,
            address=school_data.get("address"),
            phone=school_data.get("phone"),
            email=school_data.get("email"),
            principal=school_data.get("principal"),
            is_active=True
        )
        db.add(school)
        await db.flush()
        return school, "created"

    @staticmethod
    async def _generate_region_code(region_name: str, db: AsyncSession) -> str:
        """生成唯一的区域编码"""
        # 使用名称的拼音首字母+时间戳
        # 简化实现：使用名称hash的前8位
        name_hash = hashlib.md5(region_name.encode()).hexdigest()[:8]
        timestamp = str(int(datetime.now().timestamp()))[-6:]
        code = f"REG_{name_hash}_{timestamp}"
        
        # 检查是否已存在，如果存在则追加序号
        counter = 1
        while True:
            result = await db.execute(select(Region).where(Region.code == code))
            if not result.scalar_one_or_none():
                break
            code = f"REG_{name_hash}_{timestamp}_{counter}"
            counter += 1
        
        return code

    @staticmethod
    def _infer_region_level(region_name: str) -> int:
        """推断区域级别"""
        if "省" in region_name:
            return 1
        elif "市" in region_name:
            return 2
        elif "区" in region_name or "县" in region_name:
            return 3
        else:
            return 3  # 默认区级

    @staticmethod
    async def _generate_school_code(
        school_name: str,
        region_id: int,
        db: AsyncSession
    ) -> str:
        """生成唯一的学校编码"""
        # 获取区域代码
        result = await db.execute(select(Region).where(Region.id == region_id))
        region = result.scalar_one_or_none()
        region_code = region.code if region else "REG"
        
        # 使用区域代码+学校名称hash
        name_hash = hashlib.md5(school_name.encode()).hexdigest()[:6]
        timestamp = str(int(datetime.now().timestamp()))[-4:]
        code = f"{region_code}_{name_hash}_{timestamp}"
        
        # 检查是否已存在
        counter = 1
        while True:
            result = await db.execute(select(School).where(School.code == code))
            if not result.scalar_one_or_none():
                break
            code = f"{region_code}_{name_hash}_{timestamp}_{counter}"
            counter += 1
        
        return code

    @staticmethod
    async def import_schools(
        db: AsyncSession,
        records: List[Dict[str, Any]],
        auto_create_region: bool = True
    ) -> Dict[str, Any]:
        """
        批量导入学校

        Args:
            db: 数据库会话
            records: 学校记录列表
            auto_create_region: 是否自动创建不存在的区域

        Returns:
            导入结果字典
        """
        result = {
            "total": len(records),
            "success": 0,
            "failed": 0,
            "created_regions": 0,
            "created_schools": 0,
            "updated_schools": 0,
            "skipped_schools": 0,
            "errors": []
        }

        # 用于去重：同一批导入中，相同名称的区域/学校只创建一次
        region_cache: Dict[str, Region] = {}

        for record in records:
            try:
                region_name = record["region_name"]
                row_number = record.get("row_number", 0)

                # 查找或创建区域
                if region_name in region_cache:
                    region = region_cache[region_name]
                else:
                    region = await SchoolImportService.find_or_create_region(
                        db, region_name, auto_create_region
                    )
                    if not region:
                        result["errors"].append({
                            "row": row_number,
                            "field": "区域名称",
                            "message": f"区域 '{region_name}' 不存在且不允许自动创建"
                        })
                        result["failed"] += 1
                        continue
                    
                    if region.id is None:  # 新创建的
                        await db.flush()
                        result["created_regions"] += 1
                    
                    region_cache[region_name] = region

                # 查找或创建学校
                school, operation = await SchoolImportService.find_or_create_school(
                    db, record, int(region.id)  # type: ignore
                )

                if operation == "created":
                    result["created_schools"] += 1
                    result["success"] += 1
                elif operation == "updated":
                    result["updated_schools"] += 1
                    result["success"] += 1
                elif operation == "skipped":
                    result["skipped_schools"] += 1
                    result["success"] += 1
                else:
                    result["errors"].append({
                        "row": row_number,
                        "field": "学校名称",
                        "message": f"创建学校失败: {operation}"
                    })
                    result["failed"] += 1

            except IntegrityError as e:
                result["errors"].append({
                    "row": record.get("row_number", 0),
                    "field": None,
                    "message": f"数据完整性错误: {str(e)}"
                })
                result["failed"] += 1
            except Exception as e:
                result["errors"].append({
                    "row": record.get("row_number", 0),
                    "field": None,
                    "message": f"导入失败: {str(e)}"
                })
                result["failed"] += 1

        return result
