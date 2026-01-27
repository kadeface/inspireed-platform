"""
Base Excel Importer

Provides common Excel parsing functionality for all import strategies.
"""

import logging
from openpyxl import load_workbook
from typing import Any, Dict, List, Tuple, Optional
from pathlib import Path

logger = logging.getLogger(__name__)


class ParseError(Exception):
    """Excel parsing error"""
    pass


class BaseImporter:
    """Common Excel parsing functionality"""

    @staticmethod
    def normalize_header(header: Optional[str]) -> str:
        """
        Normalize header: strip whitespace, remove asterisks and markers

        Args:
            header: Raw header string from Excel

        Returns:
            Normalized header string
        """
        if not header:
            return ""
        return str(header).strip().rstrip('*').strip()

    @staticmethod
    def load_worksheet(file_path: Path):
        """
        Load Excel worksheet with read-only mode

        Args:
            file_path: Path to Excel file

        Returns:
            Tuple of (workbook, worksheet)

        Raises:
            ParseError: If file is empty or has no worksheet
        """
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ParseError("Excel文件为空或没有工作表")
        return wb, ws

    @staticmethod
    def extract_headers(ws) -> List[str]:
        """
        Extract and normalize headers from first row

        Args:
            ws: openpyxl worksheet object

        Returns:
            List of normalized header strings
        """
        headers = [cell.value for cell in ws[1]]
        return [BaseImporter.normalize_header(h) for h in headers]

    @staticmethod
    def build_column_indices(
        headers: List[str],
        column_mapping: Dict[str, str]
    ) -> Dict[str, int]:
        """
        Map column names to indices using column mapping

        Args:
            headers: List of normalized headers
            column_mapping: Dict mapping Excel column names to field names

        Returns:
            Dict mapping field names to column indices
        """
        column_indices = {}
        for col_name, field_name in column_mapping.items():
            if col_name in headers:
                column_indices[field_name] = headers.index(col_name)
        return column_indices

    @staticmethod
    def validate_required_columns(
        column_indices: Dict[str, int],
        required_fields: List[str],
        column_mapping: Dict[str, str]
    ) -> None:
        """
        Validate all required fields are present

        Args:
            column_indices: Mapped column indices
            required_fields: List of required field names
            column_mapping: Column name to field name mapping

        Raises:
            ParseError: If required columns are missing
        """
        missing_fields = [f for f in required_fields if f not in column_indices]
        if missing_fields:
            # Convert field names back to column names for error message
            field_to_col = {v: k for k, v in column_mapping.items()}
            missing_cols = [field_to_col.get(f, f) for f in missing_fields]
            raise ParseError(f"缺少必需列: {', '.join(missing_cols)}")

    @staticmethod
    def extract_row_data(
        row: Tuple,
        column_indices: Dict[str, int],
        row_number: int
    ) -> Dict[str, Any]:
        """
        Extract data from a single row

        Args:
            row: Row tuple from openpyxl
            column_indices: Field name to column index mapping
            row_number: Row number (1-indexed) for error reporting

        Returns:
            Dict with extracted field values
        """
        record = {"row_number": row_number}

        for field_name, col_idx in column_indices.items():
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None:
                    if isinstance(value, (int, float)):
                        record[field_name] = value
                    else:
                        record[field_name] = str(value).strip() if value else None
                else:
                    record[field_name] = None
            else:
                record[field_name] = None

        return record
