"""
学生考生信息批量导入服务

提供Excel文件解析、学生验证和考号映射批量导入功能
"""

import logging
import re
from typing import Any, Dict, List, Optional, Tuple
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_
from sqlalchemy.exc import IntegrityError

from app.models import (
    User,
    Exam,
    ExamNumberMapping,
    Region,
    School,
    Grade,
    Classroom,
)

logger = logging.getLogger(__name__)


class StudentImportServiceError(Exception):
    """学生导入服务错误"""
    pass


class StudentImportService:
    """学生考生信息批量导入服务"""

    # Excel列名映射（支持多种格式）
    COLUMN_MAPPING = {
        # 区域名称
        "市(区)": "region_name",
        "区域名称": "region_name",
        "区域": "region_name",
        "市": "region_name",
        "区": "region_name",
        # 学校名称
        "学校": "school_name",
        "学校名称": "school_name",
        # 学校代码
        "学校代码": "school_code",
        "代码": "school_code",
        # 姓名
        "姓名": "full_name",
        "学生姓名": "full_name",
        # 身份证号/学籍号
        "身份证号": "student_id_number",
        "学籍号": "student_id_number",
        "学生学籍号": "student_id_number",
        # 考生号
        "考生号": "exam_number",
        "考号": "exam_number",
        "考试号": "exam_number",
        "准考证号": "exam_number",
        # 班级
        "班级": "classroom_code",
        "班级编号": "classroom_code",
    }

    # 必需列
    REQUIRED_COLUMNS = ["市(区)", "学校", "姓名", "身份证号", "考生号", "班级"]

    @staticmethod
    async def parse_student_excel(file_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        解析Excel文件

        Args:
            file_path: Excel文件路径

        Returns:
            (记录列表, 错误列表)
        """
        records = []
        errors = []

        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise StudentImportServiceError("Excel文件为空或没有工作表")

            # 获取表头
            headers = [cell.value for cell in ws[1]]  # type: ignore
            if not headers:
                raise StudentImportServiceError("Excel文件为空或没有表头")

            # 标准化列名（去除空格和星号等标记）
            headers = [str(h).strip().rstrip('*').strip() if h else "" for h in headers]

            # 查找列索引
            column_indices = {}
            for col_name, field_name in StudentImportService.COLUMN_MAPPING.items():
                if col_name in headers:
                    column_indices[field_name] = headers.index(col_name)

            # 验证必需列
            required_fields = ["region_name", "school_name", "full_name", "student_id_number", "exam_number", "classroom_code"]
            missing_fields = [f for f in required_fields if f not in column_indices]
            if missing_fields:
                # 将字段名转换回中文列名
                field_to_col = {v: k for k, v in StudentImportService.COLUMN_MAPPING.items()}
                missing_cols = [field_to_col.get(f, f) for f in missing_fields]
                raise StudentImportServiceError(f"缺少必需列: {', '.join(missing_cols)}")

            # 解析数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # type: ignore
                if not any(row):  # 跳过空行
                    continue

                try:
                    record: Dict[str, Any] = {
                        "row_number": row_idx,
                    }

                    # 提取区域名称（必需）
                    region_name = row[column_indices["region_name"]]  # type: ignore
                    if not region_name:
                        errors.append({
                            "row": row_idx,
                            "field": "市(区)",
                            "message": "市(区)不能为空"
                        })
                        continue
                    record["region_name"] = str(region_name).strip()

                    # 提取学校名称（必需）
                    school_name = row[column_indices["school_name"]]  # type: ignore
                    if not school_name:
                        errors.append({
                            "row": row_idx,
                            "field": "学校",
                            "message": "学校名称不能为空"
                        })
                        continue
                    record["school_name"] = str(school_name).strip()

                    # 提取学校代码（可选）
                    if "school_code" in column_indices:
                        school_code = row[column_indices["school_code"]]  # type: ignore
                        record["school_code"] = str(school_code).strip() if school_code else None
                    else:
                        record["school_code"] = None

                    # 提取姓名（必需）
                    full_name = row[column_indices["full_name"]]  # type: ignore
                    if not full_name:
                        errors.append({
                            "row": row_idx,
                            "field": "姓名",
                            "message": "姓名不能为空"
                        })
                        continue
                    record["full_name"] = str(full_name).strip()

                    # 提取身份证号/学籍号（必需）
                    student_id_number = row[column_indices["student_id_number"]]  # type: ignore
                    if not student_id_number:
                        errors.append({
                            "row": row_idx,
                            "field": "身份证号",
                            "message": "身份证号/学籍号不能为空"
                        })
                        continue
                    record["student_id_number"] = str(student_id_number).strip()

                    # 提取考生号（必需）
                    exam_number = row[column_indices["exam_number"]]  # type: ignore
                    if not exam_number:
                        errors.append({
                            "row": row_idx,
                            "field": "考生号",
                            "message": "考生号不能为空"
                        })
                        continue
                    record["exam_number"] = str(exam_number).strip()

                    # 提取班级编号（必需）
                    classroom_code = row[column_indices["classroom_code"]]  # type: ignore
                    if not classroom_code:
                        errors.append({
                            "row": row_idx,
                            "field": "班级",
                            "message": "班级编号不能为空"
                        })
                        continue
                    record["classroom_code"] = str(classroom_code).strip()

                    records.append(record)

                except Exception as e:
                    errors.append({
                        "row": row_idx,
                        "field": None,
                        "message": f"解析行数据失败: {str(e)}"
                    })

            wb.close()
            return records, errors

        except StudentImportServiceError:
            raise
        except Exception as e:
            raise StudentImportServiceError(f"解析Excel文件失败: {str(e)}")

    @staticmethod
    def parse_classroom_code(classroom_code: str) -> Tuple[Optional[int], Optional[int]]:
        """
        解析班级编号，提取年级和班级序号

        支持格式：
        - 501 = 5年级1班 (grade_level=5, class_seq=1)
        - 1001 = 10年级1班 (grade_level=10, class_seq=1) 高一1班
        - 1203 = 12年级3班 (grade_level=12, class_seq=3) 高三3班

        Args:
            classroom_code: 班级编号字符串

        Returns:
            (grade_level, class_seq) 或 (None, None) 如果解析失败
        """
        if not classroom_code:
            return None, None

        # 去除空格
        code = str(classroom_code).strip()

        # 匹配格式：前1-2位是年级，后2位是班级序号
        # 例如：501, 0701, 1001, 1203
        # 修复：使用非贪婪匹配确保班级部分始终是2位
        match = re.match(r'^(\d{1,2})(\d{2})$', code)
        if not match:
            return None, None

        grade_part = match.group(1)
        class_part = match.group(2)

        try:
            grade_level = int(grade_part)
            class_seq = int(class_part)

            # 验证范围：年级1-12，班级序号1-99
            if 1 <= grade_level <= 12 and 1 <= class_seq <= 99:
                return grade_level, class_seq
        except ValueError:
            pass

        return None, None

    @staticmethod
    async def find_region(
        db: AsyncSession,
        region_name: str,
    ) -> Optional[Region]:
        """
        查找区域

        Args:
            db: 数据库会话
            region_name: 区域名称

        Returns:
            Region对象，如果不存在则返回None
        """
        # 1. 精确匹配
        result = await db.execute(
            select(Region).where(Region.name == region_name)
        )
        region = result.scalar_one_or_none()
        if region:
            return region

        # 2. 模糊匹配
        result = await db.execute(
            select(Region).where(Region.name.ilike(f"%{region_name}%"))
        )
        region = result.scalar_one_or_none()
        return region

    @staticmethod
    async def find_school(
        db: AsyncSession,
        school_name: str,
        region_id: Optional[int] = None,
        school_code: Optional[str] = None,
    ) -> Optional[School]:
        """
        查找学校

        Args:
            db: 数据库会话
            school_name: 学校名称
            region_id: 区域ID（可选）
            school_code: 学校代码（可选）

        Returns:
            School对象，如果不存在则返回None
        """
        # 1. 按学校代码精确匹配
        if school_code:
            result = await db.execute(
                select(School).where(School.code == school_code)
            )
            school = result.scalar_one_or_none()
            if school:
                return school

        # 2. 按名称+区域匹配
        conditions = [School.name == school_name]
        if region_id:
            conditions.append(School.region_id == region_id)

        result = await db.execute(
            select(School).where(and_(*conditions))
        )
        school = result.scalar_one_or_none()
        return school

    @staticmethod
    async def find_classroom(
        db: AsyncSession,
        classroom_code: str,
        school_id: int,
    ) -> Optional[Classroom]:
        """
        查找班级

        通过解析班级编号（如501）来查找对应的班级

        Args:
            db: 数据库会话
            classroom_code: 班级编号（如501）
            school_id: 学校ID

        Returns:
            Classroom对象，如果不存在则返回None
        """
        # 解析班级编号
        grade_level, class_seq = StudentImportService.parse_classroom_code(classroom_code)
        if not grade_level or not class_seq:
            return None

        # 查找年级
        result = await db.execute(
            select(Grade).where(Grade.level == grade_level)
        )
        grade = result.scalar_one_or_none()
        if not grade:
            return None

        # 查找班级：通过学校ID、年级ID和班级名称匹配
        # 班级名称可能是：X年级X班、X年X班等格式
        # 尝试多种匹配方式
        classroom_name_patterns = [
            f"{grade.name}{class_seq}班",  # 如：五年级1班
            f"{grade.name}第{class_seq}班",  # 如：五年级第1班
            f"{class_seq}班",  # 如：1班
        ]

        for pattern in classroom_name_patterns:
            result = await db.execute(
                select(Classroom).where(
                    and_(
                        Classroom.school_id == school_id,
                        Classroom.grade_id == grade.id,
                        or_(
                            Classroom.name == pattern,
                            Classroom.name.ilike(f"%{pattern}%"),
                            Classroom.code == classroom_code,
                        )
                    )
                )
            )
            classroom = result.scalar_one_or_none()
            if classroom:
                return classroom

        # 如果通过名称找不到，尝试通过code字段匹配
        result = await db.execute(
            select(Classroom).where(
                and_(
                    Classroom.school_id == school_id,
                    Classroom.grade_id == grade.id,
                    Classroom.code == classroom_code,
                )
            )
        )
        classroom = result.scalar_one_or_none()
        return classroom

    @staticmethod
    async def find_student(
        db: AsyncSession,
        student_id_number: str,
    ) -> Optional[User]:
        """
        查找学生

        Args:
            db: 数据库会话
            student_id_number: 学籍号/身份证号

        Returns:
            User对象，如果不存在则返回None
        """
        result = await db.execute(
            select(User).where(User.student_id_number == student_id_number)
        )
        return result.scalar_one_or_none()

    @staticmethod
    async def import_student_exam_mappings(
        db: AsyncSession,
        exam_id: int,
        records: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """
        批量导入学生考号映射

        Args:
            db: 数据库会话
            exam_id: 考试ID
            records: 学生记录列表

        Returns:
            导入结果字典
        """
        # 验证考试是否存在
        exam_result = await db.execute(
            select(Exam).where(Exam.id == exam_id)
        )
        exam = exam_result.scalar_one_or_none()
        if not exam:
            raise StudentImportServiceError(f"考试 {exam_id} 不存在")

        result = {
            "total": len(records),
            "success": 0,
            "failed": 0,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": []
        }

        # 用于去重：同一批导入中，相同名称的区域/学校只查找一次
        region_cache: Dict[str, Region] = {}
        school_cache: Dict[tuple, School] = {}  # key: (school_name, region_id)

        for record in records:
            try:
                row_number = record.get("row_number", 0)
                region_name = record["region_name"]
                school_name = record["school_name"]
                school_code = record.get("school_code")
                full_name = record["full_name"]
                student_id_number = record["student_id_number"]
                exam_number = record["exam_number"]
                classroom_code = record["classroom_code"]

                # 1. 查找区域
                if region_name in region_cache:
                    region = region_cache[region_name]
                else:
                    region = await StudentImportService.find_region(db, region_name)
                    if not region:
                        result["errors"].append({
                            "row": row_number,
                            "field": "市(区)",
                            "message": f"区域 '{region_name}' 不存在"
                        })
                        result["failed"] += 1
                        continue
                    region_cache[region_name] = region

                # 2. 查找学校
                cache_key = (school_name, region.id)
                if cache_key in school_cache:
                    school = school_cache[cache_key]
                else:
                    school = await StudentImportService.find_school(
                        db, school_name, int(region.id), school_code  # type: ignore
                    )
                    if not school:
                        result["errors"].append({
                            "row": row_number,
                            "field": "学校",
                            "message": f"学校 '{school_name}' 在区域 '{region_name}' 中不存在"
                        })
                        result["failed"] += 1
                        continue
                    school_cache[cache_key] = school

                # 3. 查找班级
                classroom = await StudentImportService.find_classroom(
                    db, classroom_code, int(school.id)  # type: ignore
                )
                if not classroom:
                    result["errors"].append({
                        "row": row_number,
                        "field": "班级",
                        "message": f"班级编号 '{classroom_code}' 在学校 '{school_name}' 中不存在或格式不正确"
                    })
                    result["failed"] += 1
                    continue

                # 4. 查找学生
                student = await StudentImportService.find_student(db, student_id_number)
                if not student:
                    result["errors"].append({
                        "row": row_number,
                        "field": "身份证号",
                        "message": f"学籍号 '{student_id_number}' 对应的学生不存在，请先创建学生账户"
                    })
                    result["failed"] += 1
                    continue

                # 5. 验证学生是否属于该班级（可选，仅警告）
                if student.classroom_id != classroom.id:  # type: ignore
                    logger.warning(
                        f"学生 {student_id_number} 的班级ID ({student.classroom_id}) "
                        f"与导入的班级ID ({classroom.id}) 不匹配"
                    )

                # 6. 检查考号映射是否已存在
                existing_mapping_result = await db.execute(
                    select(ExamNumberMapping).where(
                        and_(
                            ExamNumberMapping.exam_id == exam_id,
                            ExamNumberMapping.exam_number == exam_number
                        )
                    )
                )
                existing_mapping = existing_mapping_result.scalar_one_or_none()

                if existing_mapping:
                    # 更新现有映射
                    if existing_mapping.student_id != student.id:  # type: ignore
                        result["errors"].append({
                            "row": row_number,
                            "field": "考生号",
                            "message": f"考生号 '{exam_number}' 已被其他学生使用"
                        })
                        result["failed"] += 1
                        continue

                    # 更新冗余字段
                    existing_mapping.student_id_number = student_id_number  # type: ignore
                    existing_mapping.school_id = int(school.id)  # type: ignore
                    existing_mapping.classroom_id = int(classroom.id)  # type: ignore
                    result["updated"] += 1
                    result["success"] += 1
                else:
                    # 创建新映射
                    new_mapping = ExamNumberMapping(
                        exam_id=exam_id,
                        student_id=int(student.id),  # type: ignore
                        exam_number=exam_number,
                        student_id_number=student_id_number,
                        school_id=int(school.id),  # type: ignore
                        classroom_id=int(classroom.id),  # type: ignore
                    )
                    db.add(new_mapping)
                    result["created"] += 1
                    result["success"] += 1

            except IntegrityError as e:
                result["errors"].append({
                    "row": record.get("row_number", 0),
                    "field": None,
                    "message": f"数据完整性错误: {str(e)}"
                })
                result["failed"] += 1
                logger.error(f"导入失败（完整性错误）: {str(e)}")
            except Exception as e:
                result["errors"].append({
                    "row": record.get("row_number", 0),
                    "field": None,
                    "message": f"导入失败: {str(e)}"
                })
                result["failed"] += 1
                logger.error(f"导入失败: {str(e)}", exc_info=True)

        return result
   