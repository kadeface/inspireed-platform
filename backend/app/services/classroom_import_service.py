"""
班级批量导入服务

提供Excel文件解析、班级验证和批量导入功能
"""

import logging
import re
from typing import Any, Dict, List, Optional, Tuple
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_
from sqlalchemy.exc import IntegrityError

from app.models import (
    Region,
    School,
    Grade,
    Classroom,
)

logger = logging.getLogger(__name__)


class ClassroomImportServiceError(Exception):
    """班级导入服务错误"""
    pass


class ClassroomImportService:
    """班级批量导入服务"""

    # 县区管理端Excel列名映射（支持多种格式）
    DISTRICT_COLUMN_MAPPING = {
        # 学校名称
        "学校名称": "school_name",
        "学校": "school_name",
        "学校名称*": "school_name",
        # 学校代码
        "学校代码": "school_code",
        "代码": "school_code",
        # 年级级别
        "年级级别": "grade_level",
        "年级级别*": "grade_level",
        "级别": "grade_level",
        "level": "grade_level",
        # 年级名称
        "年级名称": "grade_name",
        "年级": "grade_name",
        "年级名称*": "grade_name",
        # 班级编号
        "班级编号": "classroom_code",
        "班级编号*": "classroom_code",
        "班级代码": "classroom_code",
        "班级编码": "classroom_code",
        # 班级名称
        "班级名称": "classroom_name",
        "班级名称*": "classroom_name",
        "班级": "classroom_name",
        # 入学年份
        "入学年份": "enrollment_year",
        "年份": "enrollment_year",
        "届别": "enrollment_year",
        # 班级容量
        "班级容量": "capacity",
        "容量": "capacity",
        "计划人数": "capacity",
        # 班级描述
        "班级描述": "description",
        "描述": "description",
        "备注": "description",
    }

    # 学校管理端Excel列名映射（简化版，不包含学校字段）
    SCHOOL_COLUMN_MAPPING = {
        # 年级级别
        "年级级别": "grade_level",
        "年级级别*": "grade_level",
        "级别": "grade_level",
        "level": "grade_level",
        # 年级名称
        "年级名称": "grade_name",
        "年级": "grade_name",
        "年级名称*": "grade_name",
        # 班级编号
        "班级编号": "classroom_code",
        "班级编号*": "classroom_code",
        "班级代码": "classroom_code",
        "班级编码": "classroom_code",
        # 班级名称
        "班级名称": "classroom_name",
        "班级名称*": "classroom_name",
        "班级": "classroom_name",
        # 入学年份
        "入学年份": "enrollment_year",
        "年份": "enrollment_year",
        "届别": "enrollment_year",
        # 班级容量
        "班级容量": "capacity",
        "容量": "capacity",
        "计划人数": "capacity",
        # 班级描述
        "班级描述": "description",
        "描述": "description",
        "备注": "description",
    }

    # 县区管理端必需列
    DISTRICT_REQUIRED_COLUMNS = ["学校名称", "年级级别", "班级编号"]

    # 学校管理端必需列
    SCHOOL_REQUIRED_COLUMNS = ["年级级别", "班级编号"]

    @staticmethod
    async def parse_classroom_excel(
        file_path: str, is_school_admin: bool = False
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        解析Excel文件

        Args:
            file_path: Excel文件路径
            is_school_admin: 是否为学校管理员（决定使用哪种列名映射）

        Returns:
            (记录列表, 错误列表)
        """
        records = []
        errors = []

        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise ClassroomImportServiceError("Excel文件为空或没有工作表")

            # 获取表头
            headers = [cell.value for cell in ws[1]]  # type: ignore
            if not headers:
                raise ClassroomImportServiceError("Excel文件为空或没有表头")

            # 标准化列名（去除空格和星号等标记）
            headers = [
                str(h).strip().rstrip("*").strip() if h else "" for h in headers
            ]

            # 选择列名映射
            column_mapping = (
                ClassroomImportService.SCHOOL_COLUMN_MAPPING
                if is_school_admin
                else ClassroomImportService.DISTRICT_COLUMN_MAPPING
            )
            required_columns = (
                ClassroomImportService.SCHOOL_REQUIRED_COLUMNS
                if is_school_admin
                else ClassroomImportService.DISTRICT_REQUIRED_COLUMNS
            )

            # 查找列索引
            column_indices: Dict[str, int] = {}
            for idx, header in enumerate(headers):
                if header in column_mapping:
                    field_name = column_mapping[header]
                    if field_name not in column_indices:  # 避免重复
                        column_indices[field_name] = idx

            # 验证必需列
            missing_columns = []
            for req_col in required_columns:
                # 检查是否有匹配的列
                found = False
                for header, field_name in column_mapping.items():
                    if req_col == header and field_name in column_indices:
                        found = True
                        break
                if not found:
                    missing_columns.append(req_col)

            if missing_columns:
                errors.append(
                    {
                        "row": 0,
                        "field": "表头",
                        "message": f"缺少必需列：{', '.join(missing_columns)}",
                    }
                )
                return records, errors

            # 解析数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # 跳过空行
                    continue

                record: Dict[str, Any] = {}
                record_errors: List[str] = []

                # 提取字段值
                for field_name, col_idx in column_indices.items():
                    if col_idx < len(row):
                        value = row[col_idx]
                        # 处理单元格值
                        if value is not None:
                            if isinstance(value, (int, float)):
                                # 如果是数字，转换类型
                                if field_name in ["grade_level", "enrollment_year", "capacity"]:
                                    record[field_name] = int(value)
                                else:
                                    record[field_name] = str(value).strip()
                            else:
                                record[field_name] = str(value).strip() if value else None
                        else:
                            record[field_name] = None
                    else:
                        record[field_name] = None

                # 验证必需字段
                if is_school_admin:
                    if not record.get("grade_level"):
                        record_errors.append("年级级别不能为空")
                    if not record.get("classroom_code"):
                        record_errors.append("班级编号不能为空")
                else:
                    if not record.get("school_name"):
                        record_errors.append("学校名称不能为空")
                    if not record.get("grade_level"):
                        record_errors.append("年级级别不能为空")
                    if not record.get("classroom_code"):
                        record_errors.append("班级编号不能为空")

                if record_errors:
                    errors.append(
                        {
                            "row": row_idx,
                            "field": "数据验证",
                            "message": "; ".join(record_errors),
                        }
                    )
                    continue

                records.append(record)

        except Exception as e:
            logger.error(f"解析Excel文件失败: {str(e)}", exc_info=True)
            errors.append(
                {
                    "row": 0,
                    "field": "文件解析",
                    "message": f"解析Excel文件失败：{str(e)}",
                }
            )

        return records, errors

    @staticmethod
    def parse_classroom_code(classroom_code: str) -> Tuple[Optional[int], Optional[int]]:
        """
        解析班级编号，提取年级级别和班级序号

        例如：
        - 701 -> (7, 1)
        - 1001 -> (10, 1)
        - 1203 -> (12, 3)

        Args:
            classroom_code: 班级编号（字符串）

        Returns:
            (年级级别, 班级序号) 或 (None, None) 如果解析失败
        """
        if not classroom_code:
            return None, None

        # 转换为字符串并去除空格
        code = str(classroom_code).strip()

        # 从右到左解析：先确定班级序号（1-2位），然后剩下的就是年级级别
        # 这样可以正确处理：701 -> (7, 1), 1001 -> (10, 1)
        # 尝试匹配1位班级序号
        match1 = re.match(r'^(\d+)(\d{1})$', code)
        if match1:
            grade_part = match1.group(1)
            class_part = match1.group(2)
            try:
                grade_level = int(grade_part)
                class_seq = int(class_part)
                if 1 <= grade_level <= 12 and 1 <= class_seq <= 9:
                    return grade_level, class_seq
            except ValueError:
                pass

        # 尝试匹配2位班级序号
        match2 = re.match(r'^(\d+)(\d{2})$', code)
        if match2:
            grade_part = match2.group(1)
            class_part = match2.group(2)
            try:
                grade_level = int(grade_part)
                class_seq = int(class_part)
                # 验证：年级1-12，班级序号10-99
                if 1 <= grade_level <= 12 and 10 <= class_seq <= 99:
                    return grade_level, class_seq
                # 如果班级序号是01-09，去掉前导0再试
                if class_part.startswith('0'):
                    class_seq = int(class_part[1])
                    if 1 <= grade_level <= 12 and 1 <= class_seq <= 9:
                        return grade_level, class_seq
            except ValueError:
                pass

        return None, None

    @staticmethod
    async def find_grade_by_level(
        db: AsyncSession, grade_level: int
    ) -> Optional[Grade]:
        """
        通过年级级别查找年级对象

        Args:
            db: 数据库会话
            grade_level: 年级级别（1-12）

        Returns:
            Grade对象，如果不存在则返回None
        """
        result = await db.execute(select(Grade).where(Grade.level == grade_level))
        return result.scalar_one_or_none()

    @staticmethod
    def generate_classroom_name(grade_name: str, class_seq: int) -> str:
        """
        根据年级名称和班级序号生成班级名称

        Args:
            grade_name: 年级名称（如"七年级"）
            class_seq: 班级序号（如1）

        Returns:
            班级名称（如"七年级1班"）
        """
        return f"{grade_name}{class_seq}班"

    @staticmethod
    async def find_school(
        db: AsyncSession,
        school_name: str,
        region_id: Optional[int] = None,
        school_code: Optional[str] = None,
    ) -> Optional[School]:
        """
        查找学校

        Args:
            db: 数据库会话
            school_name: 学校名称
            region_id: 区域ID（可选）
            school_code: 学校代码（可选，优先使用）

        Returns:
            School对象，如果不存在则返回None
        """
        # 优先级1：学校代码精确匹配
        if school_code:
            result = await db.execute(
                select(School).where(School.code == school_code.strip())
            )
            school = result.scalar_one_or_none()
            if school:
                return school

        # 优先级2：学校名称匹配
        if school_name:
            query = select(School).where(School.name == school_name.strip())
            if region_id:
                query = query.where(School.region_id == region_id)
            result = await db.execute(query)
            school = result.scalar_one_or_none()
            if school:
                return school

        # 优先级3：学校名称模糊匹配（仅当未指定区域时）
        if school_name and not region_id:
            result = await db.execute(
                select(School).where(School.name.ilike(f"%{school_name.strip()}%"))
            )
            schools = result.scalars().all()
            if len(schools) == 1:  # 只有唯一匹配时才返回
                return schools[0]

        return None

    @staticmethod
    async def import_classrooms(
        db: AsyncSession,
        records: List[Dict[str, Any]],
        school_id: Optional[int] = None,
        region_id: Optional[int] = None,
        update_existing: bool = False,
        enrollment_year: Optional[int] = None,
        capacity: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        批量导入班级

        Args:
            db: 数据库会话
            records: 记录列表
            school_id: 学校ID（学校管理员导入时必填）
            region_id: 区域ID（县区管理员导入时可选，用于学校匹配）
            update_existing: 是否更新已存在的班级
            enrollment_year: 统一设置的入学年份（如果提供，覆盖Excel中的值）
            capacity: 统一设置的班级容量（如果提供，覆盖Excel中的值）

        Returns:
            结果字典：{"total": int, "success": int, "failed": int, "created": int, "updated": int, "skipped": int, "errors": List[Dict]}
        """
        result = {
            "total": len(records),
            "success": 0,
            "failed": 0,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [],
        }

        is_school_admin = school_id is not None

        for row_number, record in enumerate(records, start=2):
            try:
                # 1. 验证并查找学校（仅县区端）
                current_school_id = school_id
                if not is_school_admin:
                    school_name = record.get("school_name")
                    school_code = record.get("school_code")
                    if not school_name:
                        result["errors"].append(
                            {
                                "row": row_number,
                                "field": "学校名称",
                                "message": "学校名称不能为空",
                            }
                        )
                        result["failed"] += 1
                        continue

                    school = await ClassroomImportService.find_school(
                        db, school_name, region_id, school_code
                    )
                    if not school:
                        result["errors"].append(
                            {
                                "row": row_number,
                                "field": "学校名称",
                                "message": f"学校 '{school_name}' 未找到",
                            }
                        )
                        result["failed"] += 1
                        continue

                    current_school_id = int(school.id)  # type: ignore

                # 2. 验证并查找年级
                grade_level = record.get("grade_level")
                if not grade_level:
                    result["errors"].append(
                        {
                            "row": row_number,
                            "field": "年级级别",
                            "message": "年级级别不能为空",
                        }
                    )
                    result["failed"] += 1
                    continue

                grade = await ClassroomImportService.find_grade_by_level(db, grade_level)
                if not grade:
                    result["errors"].append(
                        {
                            "row": row_number,
                            "field": "年级级别",
                            "message": f"年级级别 {grade_level} 不存在",
                        }
                    )
                    result["failed"] += 1
                    continue

                # 使用数据库中的年级名称（如果用户提供了年级名称，可以验证但不强制）
                grade_name = str(grade.name)  # 确保是字符串类型

                # 3. 解析班级编号
                classroom_code = record.get("classroom_code")
                if not classroom_code:
                    result["errors"].append(
                        {
                            "row": row_number,
                            "field": "班级编号",
                            "message": "班级编号不能为空",
                        }
                    )
                    result["failed"] += 1
                    continue

                parsed_grade_level, class_seq = (
                    ClassroomImportService.parse_classroom_code(classroom_code)
                )
                if not parsed_grade_level or not class_seq:
                    result["errors"].append(
                        {
                            "row": row_number,
                            "field": "班级编号",
                            "message": f"班级编号 '{classroom_code}' 格式不正确（应为：年级级别+班级序号，如701表示7年级1班）",
                        }
                    )
                    result["failed"] += 1
                    continue

                # 验证解析出的年级级别是否与提供的年级级别一致
                if parsed_grade_level != grade_level:
                    result["errors"].append(
                        {
                            "row": row_number,
                            "field": "班级编号",
                            "message": f"班级编号 '{classroom_code}' 中的年级级别 ({parsed_grade_level}) 与提供的年级级别 ({grade_level}) 不一致",
                        }
                    )
                    result["failed"] += 1
                    continue

                # 4. 生成班级名称（如果未提供）
                classroom_name = record.get("classroom_name")
                if not classroom_name:
                    classroom_name = ClassroomImportService.generate_classroom_name(
                        grade_name, class_seq
                    )

                # 5. 准备班级数据
                classroom_data = {
                    "name": classroom_name,
                    "code": str(classroom_code).strip(),
                    "school_id": current_school_id,
                    "grade_id": int(grade.id),  # type: ignore
                    "description": record.get("description"),
                    "is_active": True,
                }

                # 6. 处理入学年份（统一设置优先）
                if enrollment_year is not None:
                    classroom_data["enrollment_year"] = enrollment_year
                elif record.get("enrollment_year"):
                    classroom_data["enrollment_year"] = record["enrollment_year"]

                # 7. 处理班级容量（统一设置优先）
                if capacity is not None:
                    classroom_data["capacity"] = capacity
                elif record.get("capacity"):
                    classroom_data["capacity"] = record["capacity"]

                # 8. 检查班级是否已存在
                existing_result = await db.execute(
                    select(Classroom).where(
                        and_(
                            Classroom.school_id == current_school_id,
                            Classroom.grade_id == int(grade.id),  # type: ignore
                            or_(
                                Classroom.code == classroom_data["code"],
                                Classroom.name == classroom_data["name"],
                            ),
                        )
                    )
                )
                existing_classroom = existing_result.scalar_one_or_none()

                if existing_classroom:
                    if update_existing:
                        # 更新现有班级
                        for field, value in classroom_data.items():
                            if field not in ["school_id", "grade_id"]:  # 不更新学校ID和年级ID
                                setattr(existing_classroom, field, value)
                        await db.commit()
                        await db.refresh(existing_classroom)
                        result["updated"] += 1
                        result["success"] += 1
                    else:
                        # 跳过已存在的班级
                        result["skipped"] += 1
                        result["success"] += 1
                else:
                    # 创建新班级
                    classroom = Classroom(**classroom_data)
                    db.add(classroom)
                    await db.flush()  # 获取ID但不提交
                    result["created"] += 1
                    result["success"] += 1

            except Exception as e:
                logger.error(f"导入第{row_number}行失败: {str(e)}", exc_info=True)
                result["errors"].append(
                    {
                        "row": row_number,
                        "field": "导入",
                        "message": f"导入失败：{str(e)}",
                    }
                )
                result["failed"] += 1

        # 提交所有新创建的班级
        try:
            await db.commit()
        except IntegrityError as e:
            logger.error(f"提交班级数据失败: {str(e)}", exc_info=True)
            await db.rollback()
            result["errors"].append(
                {
                    "row": 0,
                    "field": "数据库",
                    "message": f"提交数据失败：{str(e)}",
                }
            )

        return result
