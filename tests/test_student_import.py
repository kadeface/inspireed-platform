#!/usr/bin/env python3
"""
生成学生账户导入测试 Excel 文件
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "学生导入测试"

# 表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# 表头
headers = [
    "学校名称*",
    "学校代码",
    "年级级别*",
    "班级编号*",
    "学籍号*",
    "姓名*",
    "用户名",
    "邮箱",
    "手机号",
    "性别"
]

# 写入表头
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 测试数据
test_data = [
    # 测试用例1: 完整数据
    {
        "school": "自动化测试学校",
        "school_code": "9999",
        "grade_level": 7,
        "classroom_code": "701",
        "student_id": "110101200801011234",
        "name": "张三",
        "username": "",
        "email": "",
        "phone": "13800138001",
        "gender": "男"
    },
    # 测试用例2: 最小必填数据
    {
        "school": "自动化测试学校",
        "school_code": "9999",
        "grade_level": 7,
        "classroom_code": "701",
        "student_id": "110101200802025678",
        "name": "李四",
        "username": "",
        "email": "",
        "phone": "",
        "gender": ""
    },
    # 测试用例3: 8年级
    {
        "school": "自动化测试学校",
        "school_code": "9999",
        "grade_level": 8,
        "classroom_code": "801",
        "student_id": "110101200703039999",
        "name": "王五",
        "username": "",
        "email": "",
        "phone": "13800138003",
        "gender": "女"
    },
    # 测试用例4: 高中10年级
    {
        "school": "自动化测试学校",
        "school_code": "9999",
        "grade_level": 10,
        "classroom_code": "1001",
        "student_id": "110101200805044321",
        "name": "赵六",
        "username": "",
        "email": "",
        "phone": "13800138004",
        "gender": "男"
    },
    # 测试用例5: 12年级
    {
        "school": "自动化测试学校",
        "school_code": "9999",
        "grade_level": 12,
        "classroom_code": "1203",
        "student_id": "110101200606056789",
        "name": "孙七",
        "username": "",
        "email": "",
        "phone": "13800138005",
        "gender": "女"
    },
]

# 写入数据
for row_idx, student in enumerate(test_data, start=2):
    ws.cell(row=row_idx, column=1).value = student["school"]
    ws.cell(row=row_idx, column=2).value = student["school_code"]
    ws.cell(row=row_idx, column=3).value = student["grade_level"]
    ws.cell(row=row_idx, column=4).value = student["classroom_code"]
    ws.cell(row=row_idx, column=5).value = student["student_id"]
    ws.cell(row=row_idx, column=6).value = student["name"]
    ws.cell(row=row_idx, column=7).value = student["username"]
    ws.cell(row=row_idx, column=8).value = student["email"]
    ws.cell(row=row_idx, column=9).value = student["phone"]
    ws.cell(row=row_idx, column=10).value = student["gender"]

# 调整列宽
column_widths = [15, 12, 12, 12, 20, 12, 15, 25, 15, 8]
for col, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# 保存文件
output_file = "/Users/382241106qq.com/inspireed-platform-main/test_students_import.xlsx"
wb.save(output_file)
print(f"✅ 测试文件已生成: {output_file}")
print(f"📊 包含 {len(test_data)} 条测试数据")
print("\n测试用例说明:")
print("- 用例1: 完整数据（7年级1班）")
print("- 用例2: 最小必填数据（7年级1班）")
print("- 用例3: 8年级1班")
print("- 用例4: 高一10年级1班")
print("- 用例5: 高三12年级3班")
