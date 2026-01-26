#!/usr/bin/env python3
"""
生成带考号的学生导入Excel文件
用于测试：POST /api/v1/exams/{exam_id}/students/import
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "考号导入测试"

# 表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# 表头（必需列）
headers = [
    "市(区)",
    "学校",
    "姓名",
    "身份证号",
    "考生号",
    "班级"
]

# 写入表头
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 测试数据 - 使用已导入的学生
test_data = [
    # 测试用例1: 完整数据
    {
        "region": "市直",
        "school": "自动化测试学校",
        "name": "张三",
        "student_id": "110101200801011234",
        "exam_number": "999920240101",  # 格式：学校代码(9999) + 年份(2024) + 班级(01) + 座位(01)
        "classroom": "0701"  # 使用4位格式（年级2位 + 班级2位）
    },
    # 测试用例2: 完整数据
    {
        "region": "市直",
        "school": "自动化测试学校",
        "name": "李四",
        "student_id": "110101200802025678",
        "exam_number": "999920240102",
        "classroom": "0701"
    },
    # 测试用例3: 8年级学生（班级不存在，会失败）
    {
        "region": "市直",
        "school": "自动化测试学校",
        "name": "王五",
        "student_id": "110101200703039999",
        "exam_number": "999920240801",
        "classroom": "0801"
    },
]

# 写入数据
for row_idx, student in enumerate(test_data, start=2):
    ws.cell(row=row_idx, column=1).value = student["region"]
    ws.cell(row=row_idx, column=2).value = student["school"]
    ws.cell(row=row_idx, column=3).value = student["name"]
    ws.cell(row=row_idx, column=4).value = student["student_id"]
    ws.cell(row=row_idx, column=5).value = student["exam_number"]
    ws.cell(row=row_idx, column=6).value = student["classroom"]

# 调整列宽
column_widths = [12, 20, 12, 20, 15, 10]
for col, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# 保存文件
output_file = "/Users/382241106qq.com/inspireed-platform-main/test_exam_numbers_import.xlsx"
wb.save(output_file)
print(f"✅ 带考号的测试文件已生成: {output_file}")
print(f"📊 包含 {len(test_data)} 条测试数据")
print("\n测试用例说明:")
print("- 用例1: 7年级1班 张三 (考号: 999920240101)")
print("- 用例2: 7年级1班 李四 (考号: 999920240102)")
print("- 用例3: 8年级1班 王五 (考号: 999920240801) - 预期失败（班级不存在）")
print("\n考号格式: 学校代码(4) + 年份(4) + 班级(2) + 座位(2)")
print("示例: 9999 + 2024 + 01 + 01 = 999920240101")
